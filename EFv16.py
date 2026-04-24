"""
EFv16.py — Aplicación de Análisis de Estados Financieros de CONSUCOOP
=====================================================================
Análisis financiero de cooperativas hondureñas de ahorro y crédito.

Estructura (Tablero + 4 módulos):
  - Tablero (pantalla principal): 2 paneles (Mediana / TOP 10 Anomalías)
  - F-00 Calidad de Datos (unificación de nombres)
  - F-01 Análisis de Indicadores Financieros (basado en mediana, solo Top 10)
  - F-02 Análisis de Series de Tiempo (media móvil 3 / 12 meses, YoY)
  - F-03 Detección de Anomalías (Isolation Forest + cambios abruptos mensuales)
  - F-04 Descubrimiento Causal LiNGAM (Shapiro-Wilk, método de agrupación anual)

LLM: Gemma 4 e2b local (Ollama)
Representativo central: Mediana (no se usan medias ni promedios ponderados)

Autor: Akihiro / Colaboración: Claude (Anthropic)
"""

from __future__ import annotations

import io
import re
import unicodedata
from datetime import datetime

import numpy as np
import openpyxl
import pandas as pd
import streamlit as st
from rapidfuzz import fuzz, process
from scipy import stats
from sklearn.ensemble import IsolationForest
from sklearn.preprocessing import StandardScaler

import plotly.express as px
import plotly.graph_objects as go

# Ollama
try:
    import ollama  # type: ignore
    OLLAMA_LIB_OK = True
except Exception:
    OLLAMA_LIB_OK = False

# LiNGAM
try:
    import lingam
    LINGAM_OK = True
except Exception:
    LINGAM_OK = False

# Graphviz (para el dibujo del DAG de LiNGAM en F-04)
try:
    import graphviz  # type: ignore
    GRAPHVIZ_OK = True
except Exception:
    GRAPHVIZ_OK = False


# =========================================================
# 0. Configuración de la página
# =========================================================
st.set_page_config(
    page_title="EF v16 — Análisis de Estados Financieros CONSUCOOP",
    page_icon="🏦",
    layout="wide",
)


# =========================================================
# 1. Constantes
# =========================================================
LLM_MODEL = "gemma4:e2b"  # Único modelo local Gemma 4 e2b

TYPOS = {
    "Limitadan": "Limitada",
    "Limitad,": "Limitada",
    "Limitada,": "Limitada",
    "septiembte": "septiembre",
}

STOPWORDS = {
    "cooperativa", "cooperativas",
    "de", "del", "la", "las", "los",
    "ahorro", "y", "credito", "limitada", "mixta",
}

PATTERN_MMYYYY = re.compile(r"^\d{2}-\d{4}$")
PATTERN_BG = re.compile(r"^BG\s+\d{2}-\d{4}$")
PATTERN_ER = re.compile(r"^ER\s+\d{2}-\d{4}$")
EXCLUDED_SHEETS = {"Resumen", "Hoja1", "Balance General", "Estado de Resultado"}

ASSET_HEADERS_EXACT = [
    "DISPONIBILIDADES",
    "INVERSIONES",
    "PRESTAMOS, DESCUENTOS Y NEGOCIACIONES",
    "CUENTAS Y DOCUMENTOS POR COBRAR",
    "ACTIVOS EVENTUALES",
    "PROPIEDAD, PLANTA Y EQUIPO",
    "CARGOS DIFERIDOS",
    "ACTIVOS DE INVERSION",
    "ACTIVOS INTANGIBLES",
]

INDICATORS = {
    "liquidez": {
        "label": "💧 Liquidez",
        "short": "Liquidez",
        "formula": "DISPONIBILIDADES ÷ DEPÓSITOS DE AHORRO",
        "benchmark": 0.15,
        "benchmark_label": "Referencia: ≥ 15%",
        "direction": "higher_better",
        "percent": True,
        "explain": "Liquidez disponible para responder a retiros de depósitos. "
                   "Por debajo del 15% hay riesgo de insuficiencia de fondos a corto plazo.",
    },
    "solvencia": {
        "label": "🛡️ Solvencia",
        "short": "Solvencia",
        "formula": "PATRIMONIO TOTAL ÷ ACTIVO TOTAL",
        "benchmark": 0.10,
        "benchmark_label": "Referencia: ≥ 10%",
        "direction": "higher_better",
        "percent": True,
        "explain": "Ratio de capital propio. Por debajo del 10% la capacidad de "
                   "absorción de pérdidas es marginal.",
    },
    "roa": {
        "label": "💰 Rentabilidad (ROA)",
        "short": "Rentabilidad",
        "formula": "EXCEDENTES ÷ ACTIVO TOTAL",
        "benchmark": 0.01,
        "benchmark_label": "Referencia: > 1%",
        "direction": "higher_better",
        "percent": True,
        "explain": "Rentabilidad sobre activos. Por debajo del 1% hay debilidad "
                   "en la capacidad de generar utilidades. Negativo indica pérdidas.",
    },
    "crecimiento": {
        "label": "📈 Crecimiento (YoY Activo)",
        "short": "Crecimiento",
        "formula": "Variación interanual (YoY) del ACTIVO TOTAL",
        "benchmark": 0.04,
        "benchmark_label": "Referencia: 4% (supera la inflación)",
        "direction": "higher_better",
        "percent": True,
        "explain": "Tasa de crecimiento de activos. Por debajo de la inflación "
                   "(referencia 4%) implica contracción real.",
    },
    "calidad": {
        "label": "📉 Calidad de Cartera (Morosidad)",
        "short": "Calidad de Cartera (Morosidad)",
        "formula": "(ATRASADOS + VENCIDOS + EN EJECUCIÓN) ÷ PRÉSTAMOS",
        "benchmark": 0.05,
        "benchmark_label": "Referencia: ≤ 5%",
        "direction": "lower_better",
        "percent": True,
        "explain": "Ratio de cartera morosa (= (atrasados + vencidos + en ejecución) "
                   "÷ cartera total). **Cuanto más alto, peor la calidad**. "
                   "Por encima del 5% requiere atención en gestión crediticia; "
                   "por encima del 10% es una situación grave.",
    },
}

INDICATOR_KEYS = list(INDICATORS.keys())


# =========================================================
# 2. Utilidades de unificación de nombres
# =========================================================
def extract_liquidating(name: str) -> tuple[str, bool]:
    if not isinstance(name, str):
        return name, False
    parts = re.split(r"\s*en\s+aplicaci[oó]n", name, flags=re.IGNORECASE, maxsplit=1)
    if len(parts) > 1:
        return parts[0].strip(" ,."), True
    return name, False


def apply_typos(s: str) -> str:
    for wrong, right in TYPOS.items():
        s = s.replace(wrong, right)
    return s


def normalize_text(s: str) -> str:
    if not isinstance(s, str):
        return ""
    s = apply_typos(s)
    s = unicodedata.normalize("NFKD", s)
    s = s.encode("ascii", "ignore").decode("ascii")
    s = s.lower()
    s = re.sub(r"[^\w\s]", " ", s)
    s = re.sub(r"\s+", " ", s).strip()
    return s


def core_name(name: str) -> str:
    body, _ = extract_liquidating(name)
    normalized = normalize_text(body)
    tokens = [t for t in normalized.split() if t not in STOPWORDS]
    return " ".join(tokens)


def cluster_names(
    names: list[str],
    threshold_auto: int = 85,
    threshold_review: int = 70,
) -> tuple[list[dict], list[tuple[str, str, float]]]:
    buckets: dict[str, list[str]] = {}
    review: list[tuple[str, str, float]] = []
    liq_map: dict[str, bool] = {}

    sorted_names = sorted(set(names), key=lambda x: -len(x))
    for name in sorted_names:
        _, is_liq = extract_liquidating(name)
        liq_map[name] = is_liq
        core = core_name(name)
        if not core:
            buckets.setdefault("__unknown__", []).append(name)
            continue
        if buckets:
            match = process.extractOne(
                core, list(buckets.keys()), scorer=fuzz.token_sort_ratio
            )
            if match is not None:
                matched_core, score, _ = match
                if score >= threshold_auto:
                    buckets[matched_core].append(name)
                    continue
                elif score >= threshold_review:
                    review.append((name, matched_core, float(score)))
                    buckets[core] = [name]
                    continue
        buckets[core] = [name]

    clusters: list[dict] = []
    for cid, members in buckets.items():
        representative = sorted(members, key=lambda x: -len(x))[0]
        any_liq = any(liq_map.get(m, False) for m in members)
        all_liq = all(liq_map.get(m, False) for m in members)
        clusters.append({
            "cluster_id": cid,
            "representative": representative,
            "members": members,
            "member_count": len(members),
            "is_liquidating": all_liq,
            "has_liquidating_variant": any_liq,
        })
    return clusters, review


def build_name_mapping(clusters: list[dict]) -> dict[str, dict]:
    mapping: dict[str, dict] = {}
    for c in clusters:
        for m in c["members"]:
            _, is_liq = extract_liquidating(m)
            mapping[m] = {
                "representative": c["representative"],
                "cluster_is_liquidating": c["is_liquidating"],
                "member_is_liquidating": is_liq,
            }
    return mapping


# =========================================================
# 3. Lectura de Excel
# =========================================================
def detect_sheet_format(sheet_name: str) -> str:
    if sheet_name in EXCLUDED_SHEETS:
        return "excluded"
    if PATTERN_MMYYYY.match(sheet_name):
        return "MMYYYY"
    if PATTERN_BG.match(sheet_name):
        return "BG"
    if PATTERN_ER.match(sheet_name):
        return "ER"
    return "unknown"


def sheet_to_period(sheet_name: str, fmt: str) -> tuple[int | None, int | None]:
    try:
        if fmt == "MMYYYY":
            mm, yyyy = sheet_name.split("-")
            return int(yyyy), int(mm)
        elif fmt in {"BG", "ER"}:
            _, rest = sheet_name.split(" ", 1)
            mm, yyyy = rest.split("-")
            return int(yyyy), int(mm)
    except Exception:
        pass
    return None, None


def _safe_num(v) -> float | None:
    if v is None:
        return None
    if isinstance(v, (int, float)):
        return float(v)
    try:
        return float(str(v).replace(",", "").strip())
    except Exception:
        return None


def _norm_header(h) -> str:
    if h is None:
        return ""
    s = str(h).strip()
    s = unicodedata.normalize("NFKD", s)
    s = s.encode("ascii", "ignore").decode("ascii")
    return s.upper().strip()


def _build_header_map(header_row) -> dict[str, list[int]]:
    m: dict[str, list[int]] = {}
    for j, h in enumerate(header_row):
        nh = _norm_header(h)
        if not nh:
            continue
        m.setdefault(nh, []).append(j)
    return m


def _first_by_header(row, hmap: dict[str, list[int]], key: str) -> float | None:
    k = _norm_header(key)
    indices = hmap.get(k, [])
    if not indices:
        return None
    idx = indices[0]
    if idx < len(row):
        return _safe_num(row[idx])
    return None


def _sum_by_prefix(row, hmap: dict[str, list[int]], prefix: str) -> float | None:
    pfx = _norm_header(prefix)
    total = 0.0
    found = False
    for k, indices in hmap.items():
        if k == pfx or k.startswith(pfx + " ") or k.startswith(pfx + ","):
            for idx in indices:
                if idx < len(row):
                    v = _safe_num(row[idx])
                    if v is not None:
                        total += v
                        found = True
    return total if found else None


# Encabezados a ignorar en el formato largo de datos brutos (columnas meta)
_EXCLUDED_RAW_HEADERS = {"COOPERATIVA", "MES", "ANO", "AO", ""}


@st.cache_data(show_spinner=False)
def load_excel(file_bytes: bytes) -> tuple[list[str], list[dict], pd.DataFrame, pd.DataFrame]:
    """Retorna, además de los DataFrames originales, df_raw_all en formato largo
    con todas las variables brutas.
    """
    wb = openpyxl.load_workbook(io.BytesIO(file_bytes), data_only=True, read_only=True)

    sheet_info: list[dict] = []
    all_names: set[str] = set()
    records: list[dict] = []
    records_all: list[dict] = []  # Formato largo (todas las variables brutas)

    for sname in wb.sheetnames:
        fmt = detect_sheet_format(sname)
        ws = wb[sname]
        rows_iter = list(ws.iter_rows(values_only=True))
        row_count = len(rows_iter)
        col_count = ws.max_column or 0
        sheet_info.append({"sheet": sname, "format": fmt, "rows": row_count, "cols": col_count})

        if fmt in {"excluded", "unknown"}:
            continue
        if row_count < 2:
            continue

        header_row = rows_iter[0]
        hmap = _build_header_map(header_row)
        year, month = sheet_to_period(sname, fmt)

        for row in rows_iter[1:]:
            coop = row[0] if len(row) > 0 else None
            if not coop or not isinstance(coop, str) or not coop.strip():
                continue
            coop = coop.strip()
            all_names.add(coop)

            rec: dict = {
                "coop_original": coop,
                "year": year,
                "month": month,
                "sheet_format": fmt,
                "sheet": sname,
                "activos_totales": None,
                "disponibilidades": None,
                "prestamos_total": None,
                "bad_loans": None,
                "depositos_ahorro": None,
                "patrimonio_total": None,
                "excedentes": None,
            }

            if fmt in {"MMYYYY", "BG"}:
                rec["disponibilidades"] = _first_by_header(row, hmap, "DISPONIBILIDADES")
                rec["prestamos_total"] = _first_by_header(
                    row, hmap, "PRESTAMOS, DESCUENTOS Y NEGOCIACIONES"
                )
                atrasados = _sum_by_prefix(row, hmap, "ATRASADOS")
                vencidos = _sum_by_prefix(row, hmap, "VENCIDOS")
                ejecucion = _sum_by_prefix(row, hmap, "EN EJECUCION JUDICIAL")
                parts = [x for x in (atrasados, vencidos, ejecucion) if x is not None]
                rec["bad_loans"] = sum(parts) if parts else None
                rec["depositos_ahorro"] = _sum_by_prefix(row, hmap, "DEPOSITOS DE AHORRO")
                patrimonio_parent = _first_by_header(row, hmap, "PATRIMONIO")
                if patrimonio_parent is None:
                    pp = _first_by_header(row, hmap, "PATRIMONIO PRIMARIO") or 0
                    pc = _first_by_header(row, hmap, "PATRIMONIO COMPLEMENTARIO") or 0
                    rec["patrimonio_total"] = (pp + pc) if (pp or pc) else None
                else:
                    rec["patrimonio_total"] = patrimonio_parent
                at = _first_by_header(row, hmap, "ACTIVOS TOTALES")
                if at is None:
                    total = 0.0
                    any_found = False
                    for k in ASSET_HEADERS_EXACT:
                        v = _first_by_header(row, hmap, k)
                        if v is not None:
                            total += v
                            any_found = True
                    rec["activos_totales"] = total if any_found else None
                else:
                    rec["activos_totales"] = at
                if fmt == "MMYYYY":
                    rec["excedentes"] = _first_by_header(
                        row, hmap, "EXCEDENTES O PERDIDAS DEL PERIODO"
                    )
            elif fmt == "ER":
                rec["excedentes"] = _first_by_header(
                    row, hmap, "EXCEDENTES O PERDIDAS DEL PERIODO"
                )

            records.append(rec)

            # Formato largo: se recolectan todas las variables brutas (columnas del mismo nombre se suman)
            if fmt in {"MMYYYY", "BG", "ER"}:
                for nh, indices in hmap.items():
                    if not nh or nh in _EXCLUDED_RAW_HEADERS:
                        continue
                    total = 0.0
                    found_any = False
                    for idx in indices:
                        if idx < len(row):
                            v = _safe_num(row[idx])
                            if v is not None:
                                total += v
                                found_any = True
                    if found_any:
                        records_all.append({
                            "coop_original": coop,
                            "year": year,
                            "month": month,
                            "sheet_format": fmt,
                            "sheet": sname,
                            "variable": nh,
                            "value": total,
                        })
                # MMYYYY: no existe fila padre PATRIMONIO → se sintetiza como PRIMARIO + COMPLEMENTARIO
                if fmt == "MMYYYY":
                    pp = _first_by_header(row, hmap, "PATRIMONIO PRIMARIO")
                    pc = _first_by_header(row, hmap, "PATRIMONIO COMPLEMENTARIO")
                    if pp is not None or pc is not None:
                        records_all.append({
                            "coop_original": coop,
                            "year": year,
                            "month": month,
                            "sheet_format": fmt,
                            "sheet": sname,
                            "variable": "PATRIMONIO",
                            "value": (pp or 0.0) + (pc or 0.0),
                        })
                    # ACTIVOS TOTALES tampoco tiene fila padre → se sintetiza
                    at = _first_by_header(row, hmap, "ACTIVOS TOTALES")
                    if at is None:
                        t = 0.0
                        found = False
                        for k in ASSET_HEADERS_EXACT:
                            v = _first_by_header(row, hmap, k)
                            if v is not None:
                                t += v
                                found = True
                        if found:
                            records_all.append({
                                "coop_original": coop,
                                "year": year,
                                "month": month,
                                "sheet_format": fmt,
                                "sheet": sname,
                                "variable": "ACTIVOS TOTALES",
                                "value": t,
                            })

    df_raw = pd.DataFrame(records)
    df_raw_all = pd.DataFrame(records_all)
    return sorted(all_names), sheet_info, df_raw, df_raw_all


# =========================================================
# 4. Cálculo de indicadores financieros
# =========================================================
def compute_indicators(df_raw: pd.DataFrame, name_mapping: dict[str, dict]) -> pd.DataFrame:
    if df_raw.empty:
        return pd.DataFrame()

    df = df_raw.copy()
    df["coop"] = df["coop_original"].map(
        lambda x: name_mapping.get(x, {}).get("representative", x)
    )
    df["is_liquidating"] = df["coop_original"].map(
        lambda x: name_mapping.get(x, {}).get("cluster_is_liquidating", False)
    )

    df_mm = df[df["sheet_format"] == "MMYYYY"].copy()
    df_bg = df[df["sheet_format"] == "BG"].copy()
    df_er = df[df["sheet_format"] == "ER"].copy()

    if not df_bg.empty and not df_er.empty:
        er_slim = df_er.groupby(["coop", "year", "month"], as_index=False)["excedentes"].first()
        if "excedentes" in df_bg.columns:
            df_bg = df_bg.drop(columns=["excedentes"])
        df_bg = df_bg.merge(er_slim, on=["coop", "year", "month"], how="left")

    common_cols = [
        "coop", "coop_original", "year", "month", "is_liquidating", "sheet_format",
        "activos_totales", "disponibilidades", "prestamos_total",
        "bad_loans", "depositos_ahorro", "patrimonio_total", "excedentes",
    ]
    frames = []
    for frame in (df_mm, df_bg):
        if frame.empty:
            continue
        missing = [c for c in common_cols if c not in frame.columns]
        for c in missing:
            frame[c] = None
        frames.append(frame[common_cols])
    if not frames:
        return pd.DataFrame()
    combined = pd.concat(frames, ignore_index=True)

    combined["_fmt_priority"] = combined["sheet_format"].map({"MMYYYY": 0, "BG": 1}).fillna(2)
    combined = (
        combined.sort_values(["coop", "year", "month", "_fmt_priority"])
        .drop_duplicates(subset=["coop", "year", "month"], keep="first")
        .drop(columns=["_fmt_priority"])
        .reset_index(drop=True)
    )

    def safe_div(n, d):
        n = pd.to_numeric(n, errors="coerce")
        d = pd.to_numeric(d, errors="coerce")
        with np.errstate(divide="ignore", invalid="ignore"):
            out = np.where((d.isna()) | (d == 0), np.nan, n / d)
        return pd.Series(out, index=n.index)

    combined["liquidez"] = safe_div(combined["disponibilidades"], combined["depositos_ahorro"])
    combined["solvencia"] = safe_div(combined["patrimonio_total"], combined["activos_totales"])
    combined["roa"] = safe_div(combined["excedentes"], combined["activos_totales"])
    combined["calidad"] = safe_div(combined["bad_loans"], combined["prestamos_total"])

    combined = combined.sort_values(["coop", "year", "month"]).reset_index(drop=True)
    combined["period"] = combined["year"].astype("Int64") * 12 + combined["month"].astype("Int64")
    prev = combined[["coop", "period", "activos_totales"]].rename(
        columns={"period": "period_prev", "activos_totales": "activos_yoy_prev"}
    )
    prev["period_prev"] = prev["period_prev"] + 12
    combined = combined.merge(
        prev, left_on=["coop", "period"], right_on=["coop", "period_prev"], how="left",
    ).drop(columns=["period_prev"])
    combined["crecimiento"] = safe_div(
        combined["activos_totales"] - combined["activos_yoy_prev"],
        combined["activos_yoy_prev"],
    )

    combined["date"] = pd.to_datetime(
        combined["year"].astype("Int64").astype(str) + "-"
        + combined["month"].astype("Int64").astype(str).str.zfill(2) + "-01",
        errors="coerce",
    )

    return combined


# =========================================================
# 4.5 Procesamiento del formato largo de variables brutas
# =========================================================
def compute_raw_all(
    df_raw_all: pd.DataFrame, name_mapping: dict[str, dict],
) -> pd.DataFrame:
    """Agrega nombre representativo, fecha y bandera de liquidación al formato largo
    de variables brutas, y resuelve duplicados (MMYYYY vs BG/ER) por prioridad.
    """
    if df_raw_all.empty:
        return pd.DataFrame()
    df = df_raw_all.copy()
    df["coop"] = df["coop_original"].map(
        lambda x: name_mapping.get(x, {}).get("representative", x)
    )
    df["is_liquidating"] = df["coop_original"].map(
        lambda x: name_mapping.get(x, {}).get("cluster_is_liquidating", False)
    )
    df["date"] = pd.to_datetime(
        df["year"].astype("Int64").astype(str) + "-"
        + df["month"].astype("Int64").astype(str).str.zfill(2) + "-01",
        errors="coerce",
    )
    # Prioridad: MMYYYY (hoja unificada) > BG/ER
    df["_fmt_priority"] = df["sheet_format"].map(
        {"MMYYYY": 0, "BG": 1, "ER": 1}
    ).fillna(2)
    df = (
        df.sort_values(["coop", "date", "variable", "_fmt_priority"])
        .drop_duplicates(subset=["coop", "date", "variable"], keep="first")
        .drop(columns=["_fmt_priority"])
        .reset_index(drop=True)
    )
    return df


def get_recent_all_raw_data(
    df_raw_all_clean: pd.DataFrame,
    coop: str,
    target_date: pd.Timestamp,
    n_months: int = 3,
) -> pd.DataFrame:
    """Retorna **todas las variables brutas** de los últimos n meses de la cooperativa
    indicada, en formato tabla (filas = variable, columnas = mes), con columna
    adicional de variación mensual (%).
    """
    if df_raw_all_clean.empty:
        return pd.DataFrame()
    df = df_raw_all_clean[df_raw_all_clean["coop"] == coop].copy()
    df = df[df["date"] <= target_date].sort_values("date")
    dates = sorted(df["date"].dropna().unique())
    if not dates:
        return pd.DataFrame()
    dates = dates[-n_months:]
    df = df[df["date"].isin(dates)]
    if df.empty:
        return pd.DataFrame()
    pivot = df.pivot_table(
        index="variable", columns="date", values="value", aggfunc="first",
    )
    pivot = pivot[sorted(pivot.columns)]
    # Renombrar columnas de fecha al formato YYYY-MM
    pivot = pivot.rename(columns=lambda d: pd.Timestamp(d).strftime("%Y-%m"))

    # Calcular variación mensual % (mes final vs mes anterior)
    pct_display = None
    if pivot.shape[1] >= 2:
        last_col = pivot.columns[-1]
        prev_col = pivot.columns[-2]
        with np.errstate(divide="ignore", invalid="ignore"):
            pct = np.where(
                pivot[prev_col].notna() & pivot[last_col].notna()
                & (pivot[prev_col].abs() > 0),
                (pivot[last_col] - pivot[prev_col]) / pivot[prev_col].abs() * 100,
                np.nan,
            )
        pct_display = [f"{v:+.1f}%" if pd.notna(v) else "—" for v in pct]

    # Formatear valores en miles de HNL para mayor legibilidad
    for col in list(pivot.columns):
        pivot[col] = pivot[col].apply(
            lambda v: f"{v/1000:,.0f}" if pd.notna(v) else "—"
        )
    if pct_display is not None:
        pivot["Var. Mensual %"] = pct_display

    pivot = pivot.reset_index().rename(columns={"variable": "Variable"})
    return pivot


# Los valores que se redondean en miles de HNL y aparecen como '0' o '-0' en
# pantalla (|value| < 500 HNL) tienen alta probabilidad de ser sustitutos de
# valores faltantes o errores de redondeo, e inflan la variación % mensual.
# Por eso se excluyen los valores con |value| < 1000 HNL.
RAW_PCT_MIN_ABS_VALUE_HNL = 1000.0  # Valores menores a 1 mil HNL quedan fuera


def detect_raw_pct_changes(
    df_raw_all_clean: pd.DataFrame,
    target_date: pd.Timestamp | None = None,
    top_k: int = 20,
    min_abs_value: float = RAW_PCT_MIN_ABS_VALUE_HNL,
) -> pd.DataFrame:
    """Top k variables brutas con mayor variación mensual (%) en valor absoluto.

    Al comparar por % en vez de valor absoluto se evita la dependencia de la
    escala de cada variable y se detectan los 'movimientos amplios'
    independientemente de su magnitud.

    Se excluyen filas donde el mes actual o el mes anterior tengan valor 0,
    faltante, o un valor absoluto inferior a min_abs_value (por defecto 1 mil HNL),
    ya que estos se visualizarán como '0' o '-0' en pantalla al redondear a miles.
    """
    if df_raw_all_clean.empty:
        return pd.DataFrame()
    df = df_raw_all_clean[~df_raw_all_clean["is_liquidating"]].copy()
    if df.empty or "date" not in df.columns:
        return pd.DataFrame()
    if target_date is None:
        target_date = df["date"].max()
    prev_date = (pd.Timestamp(target_date) - pd.DateOffset(months=1)).normalize()
    prev_date = prev_date.replace(day=1)

    cur = df[df["date"] == target_date][["coop", "variable", "value"]].rename(
        columns={"value": "value_cur"}
    )
    prv = df[df["date"] == prev_date][["coop", "variable", "value"]].rename(
        columns={"value": "value_prev"}
    )
    merged = cur.merge(prv, on=["coop", "variable"], how="inner")
    if merged.empty:
        return pd.DataFrame()
    # Exclusión de filas con valores faltantes, cero o menores al umbral mínimo
    thr = float(min_abs_value)
    merged = merged[
        merged["value_prev"].notna() & merged["value_cur"].notna()
        & (merged["value_prev"] != 0) & (merged["value_cur"] != 0)
        & (merged["value_prev"].abs() >= thr)
        & (merged["value_cur"].abs() >= thr)
    ]
    if merged.empty:
        return pd.DataFrame()
    merged["pct_change"] = (
        (merged["value_cur"] - merged["value_prev"]) / merged["value_prev"].abs() * 100
    )
    merged["abs_pct_change"] = merged["pct_change"].abs()
    top = merged.nlargest(top_k, "abs_pct_change").reset_index(drop=True)
    top["Dirección"] = top["pct_change"].apply(lambda x: "↑ Aumento" if x > 0 else "↓ Disminución")
    return top


def run_lingam_raw(
    df_raw_all_clean: pd.DataFrame,
    selected_vars: list[str],
    target_date: pd.Timestamp | None = None,
) -> tuple[np.ndarray | None, list[str], pd.DataFrame | None]:
    """Ejecuta DirectLiNGAM sobre las variables brutas seleccionadas.
    Se excluyen los valores faltantes en la cooperativa × variable del mes objetivo
    y se ejecuta si hay al menos 10 muestras disponibles.
    """
    if not LINGAM_OK or len(selected_vars) < 2 or df_raw_all_clean.empty:
        return None, [], None
    df = df_raw_all_clean[~df_raw_all_clean["is_liquidating"]].copy()
    if df.empty or "date" not in df.columns:
        return None, [], None
    if target_date is None:
        target_date = df["date"].max()
    df_t = df[df["date"] == target_date]
    df_t = df_t[df_t["variable"].isin(selected_vars)]
    if df_t.empty:
        return None, [], None
    wide = df_t.pivot_table(
        index="coop", columns="variable", values="value", aggfunc="first",
    )
    avail = [v for v in selected_vars if v in wide.columns]
    if len(avail) < 2:
        return None, avail, None
    wide = wide[avail].dropna()
    if len(wide) < 10:
        return None, avail, None

    scaler = StandardScaler()
    X_scaled = scaler.fit_transform(wide.values)
    try:
        model = lingam.DirectLiNGAM()
        model.fit(X_scaled)
        adj = model.adjacency_matrix_
        return adj, avail, wide
    except Exception as e:
        st.warning(f"Error en cálculo de LiNGAM: {e}")
        return None, avail, None


# =========================================================
# 5. Invocación de Gemma 4 e2b local
# =========================================================
def call_gemma(
    prompt: str,
    system: str | None = None,
    host: str | None = None,
    timeout: int = 180,
) -> str:
    """Invocación de Gemma 4 e2b local (Ollama). Sin UI de selección de modelo."""
    messages: list[dict] = []
    if system:
        messages.append({"role": "system", "content": system})
    messages.append({"role": "user", "content": prompt})

    if OLLAMA_LIB_OK:
        try:
            client_kwargs = {"host": host} if host else {}
            client = ollama.Client(**client_kwargs) if client_kwargs else ollama
            resp = client.chat(
                model=LLM_MODEL,
                messages=messages,
                options={"temperature": 0.3},
                stream=False,
            )
            return resp["message"]["content"].strip()
        except Exception as e:
            return (
                f"⚠️ Falló la invocación de Ollama: {e}\n\n"
                f"Verifique:\n"
                f"1. Que `ollama serve` esté en ejecución\n"
                f"2. Que se haya ejecutado `ollama pull {LLM_MODEL}`\n"
                f"3. Que la URL del host sea correcta"
            )

    try:
        import requests
    except ImportError:
        return "⚠️ No están disponibles ni la librería ollama ni requests. Ejecute `pip install ollama`."

    base_host = host or "http://localhost:11434"
    try:
        r = requests.post(
            f"{base_host}/api/chat",
            json={
                "model": LLM_MODEL,
                "messages": messages,
                "options": {"temperature": 0.3},
                "stream": False,
            },
            timeout=timeout,
        )
        r.raise_for_status()
        j = r.json()
        return (j.get("message", {}) or {}).get("content", "").strip()
    except Exception as e:
        return f"⚠️ Falló la invocación HTTP de Ollama: {e}"


# =========================================================
# 6. Utilidades comunes
# =========================================================
def _fmt_value(v: float, percent: bool) -> str:
    if v is None or (isinstance(v, float) and np.isnan(v)):
        return "N/A"
    return f"{v*100:.1f}%" if percent else f"{v:.3f}"


def _trend_arrow(curr: float, prev: float, lower_better: bool = False) -> str:
    """Flecha de variación mensual. Si lower_better es True, no se aplica la lógica
    de 'bajar = mejorar' para que la interpretación quede del lado del usuario.
    """
    if curr is None or prev is None or np.isnan(curr) or np.isnan(prev):
        return "→"
    diff = curr - prev
    # Se muestra solo la dirección. La interpretación de mejora/deterioro queda al usuario.
    if abs(diff) < 1e-6:
        return "→"
    return "↑" if diff > 0 else "↓"


def _filter_active(df: pd.DataFrame) -> pd.DataFrame:
    """Retorna los datos excluyendo las cooperativas en liquidación."""
    if "is_liquidating" not in df.columns:
        return df
    return df[~df["is_liquidating"]].copy()


def get_recent_raw_data(
    df_ind: pd.DataFrame,
    coop: str,
    target_date: pd.Timestamp,
    n_months: int = 3,
) -> pd.DataFrame:
    """Retorna los últimos n meses de la cooperativa con los datos brutos
    (valores financieros) más los indicadores calculados, en formato tabla.

    Ayuda a distinguir si una anomalía detectada corresponde a un error de
    registro o a un problema real de gestión: los saltos de magnitud o
    variaciones bruscas se visualizan de un vistazo.
    """
    df = df_ind[df_ind["coop"] == coop].copy()
    df = df[df["date"] <= target_date].sort_values("date").tail(n_months)
    if df.empty:
        return pd.DataFrame()

    # Formato: Mes / Valores brutos (HNL) / Indicadores (%)
    result = pd.DataFrame({"Mes": df["date"].dt.strftime("%Y-%m").values})

    # Valores brutos — se redondean a miles de Lempiras para mayor legibilidad
    raw_cols = {
        "Activo Total (miles HNL)": "activos_totales",
        "Disponibilidades (miles HNL)": "disponibilidades",
        "Préstamos (miles HNL)": "prestamos_total",
        "Cartera Morosa (miles HNL)": "bad_loans",
        "Depósitos (miles HNL)": "depositos_ahorro",
        "Patrimonio (miles HNL)": "patrimonio_total",
        "Excedentes (miles HNL)": "excedentes",
    }
    for label, col in raw_cols.items():
        if col in df.columns:
            vals = pd.to_numeric(df[col], errors="coerce") / 1000.0
            result[label] = [f"{v:,.0f}" if pd.notna(v) else "—" for v in vals]

    # Indicadores (%)
    for k in INDICATOR_KEYS:
        label = INDICATORS[k]["short"]
        vals = df[k].values
        result[f"{label} (%)"] = [
            f"{v*100:.1f}%" if pd.notna(v) else "—" for v in vals
        ]

    return result


# =========================================================
# 7. F-01 Indicadores Financieros (basado en mediana, solo Top 10)
# =========================================================
def build_sector_median_chart(
    df_ind: pd.DataFrame, ind_key: str
) -> tuple[go.Figure, pd.DataFrame]:
    """Evolución de la mediana (sin medias ni promedios ponderados)."""
    cfg = INDICATORS[ind_key]
    df = _filter_active(df_ind).dropna(subset=[ind_key, "date"])
    if df.empty:
        return go.Figure(), pd.DataFrame()

    grouped = df.groupby("date")[ind_key].median().reset_index()
    series_name = "Mediana sectorial"

    fig = go.Figure()
    fig.add_trace(go.Scatter(
        x=grouped["date"], y=grouped[ind_key],
        mode="lines+markers", name=series_name,
        line=dict(width=3, color="#1F4E79"),
    ))
    fig.add_hline(
        y=cfg["benchmark"],
        line=dict(color="red" if cfg["direction"] == "higher_better" else "orange", dash="dash"),
        annotation_text=cfg["benchmark_label"],
        annotation_position="top right",
    )
    fig.update_layout(
        title=f"Evolución de {cfg['short']} (mediana)",
        yaxis_tickformat=".0%" if cfg["percent"] else ".3f",
        xaxis_title="Mes", yaxis_title=cfg["short"],
        height=360, margin=dict(l=40, r=20, t=50, b=40),
    )
    return fig, grouped


# =========================================================
# 8. F-02 Series de Tiempo (media móvil 3 / 12 meses, YoY)
# =========================================================
def build_time_series_chart(
    df_ind: pd.DataFrame,
    selected_coops: list[str],
    selected_indicators: list[str],
    ma_window: int,
    show_yoy: bool,
    treat_zero_as_nan: bool,
) -> go.Figure:
    """Gráfico de series de tiempo: media móvil 3/12 meses, YoY opcional."""
    df = df_ind.copy()
    df = df[df["coop"].isin(selected_coops)].dropna(subset=["date"])
    if df.empty:
        return go.Figure()

    fig = go.Figure()
    color_palette = px.colors.qualitative.Plotly

    color_idx = 0
    for coop in selected_coops:
        df_c = df[df["coop"] == coop].sort_values("date")
        for ind in selected_indicators:
            cfg = INDICATORS[ind]
            series = df_c[["date", ind]].dropna(subset=[ind])
            if treat_zero_as_nan:
                series = series[series[ind] != 0]
            if series.empty:
                continue

            # Media móvil
            series = series.set_index("date")[ind]
            if ma_window > 1:
                series_ma = series.rolling(window=ma_window, min_periods=1).mean()
                ma_label = f" (MM {ma_window}m)"
            else:
                series_ma = series
                ma_label = ""

            color = color_palette[color_idx % len(color_palette)]
            fig.add_trace(go.Scatter(
                x=series_ma.index, y=series_ma.values,
                mode="lines+markers",
                name=f"{coop} - {cfg['short']}{ma_label}",
                line=dict(color=color, width=2),
            ))

            # Superposición YoY (opcional)
            if show_yoy and len(series_ma) >= 13:
                yoy_pct = (series_ma / series_ma.shift(12) - 1) * 100
                yoy_pct = yoy_pct.dropna()
                if not yoy_pct.empty:
                    fig.add_trace(go.Scatter(
                        x=yoy_pct.index, y=yoy_pct.values,
                        mode="lines", line=dict(color=color, dash="dot", width=1),
                        name=f"{coop} - {cfg['short']} YoY (%)",
                        yaxis="y2",
                    ))
            color_idx += 1

    layout_kwargs = dict(
        title="Evolución en el tiempo",
        xaxis_title="Mes",
        height=520,
        margin=dict(l=40, r=60, t=50, b=40),
        legend=dict(orientation="h", yanchor="bottom", y=-0.3, xanchor="left", x=0),
    )
    if show_yoy:
        layout_kwargs["yaxis"] = dict(title="Valor del indicador")
        layout_kwargs["yaxis2"] = dict(
            title="YoY (%)", overlaying="y", side="right", showgrid=False,
        )
    fig.update_layout(**layout_kwargs)
    return fig


# =========================================================
# 9. F-03 Detección de Anomalías (Isolation Forest + cambios abruptos)
# =========================================================
def detect_anomalies_isoforest(
    df_ind: pd.DataFrame, target_date: pd.Timestamp | None = None,
    contamination: float = 0.15,
) -> pd.DataFrame:
    """Detección de anomalía global (corte transversal) mediante Isolation Forest.

    Se aplica al conjunto cooperativa × todos los indicadores del mes más reciente
    tras su estandarización. Se retorna en orden descendente de score de anomalía.
    """
    df = _filter_active(df_ind).dropna(subset=["date"])
    if df.empty:
        return pd.DataFrame()

    if target_date is None:
        target_date = df["date"].max()

    df_t = df[df["date"] == target_date].copy()
    df_t = df_t.drop_duplicates(subset=["coop"], keep="first")
    if df_t.empty or len(df_t) < 5:
        return pd.DataFrame()

    # Se usan todos los indicadores
    X = df_t[INDICATOR_KEYS].copy()
    # Imputación de faltantes con la mediana
    X = X.fillna(X.median(numeric_only=True))
    # Si aún hay faltantes, se llenan con 0
    X = X.fillna(0)

    if len(X) < 5:
        return pd.DataFrame()

    scaler = StandardScaler()
    X_scaled = scaler.fit_transform(X)

    iso = IsolationForest(
        contamination=contamination, random_state=42, n_estimators=100,
    )
    iso.fit(X_scaled)
    scores = -iso.score_samples(X_scaled)  # A mayor valor, más anómalo
    is_anomaly = iso.predict(X_scaled) == -1

    # Contribución: valor absoluto estandarizado de cada indicador
    contributions = pd.DataFrame(
        np.abs(X_scaled), columns=INDICATOR_KEYS, index=df_t.index,
    )

    result = df_t[["coop", "date"]].copy()
    result["anomaly_score"] = scores
    result["is_anomaly"] = is_anomaly
    # Preocupación principal = indicador con mayor contribución
    result["top_concern"] = contributions.idxmax(axis=1).map(
        lambda k: INDICATORS[k]["short"]
    )
    # Contribución de cada indicador
    for k in INDICATOR_KEYS:
        result[f"contrib_{k}"] = contributions[k].values

    return result.sort_values("anomaly_score", ascending=False).reset_index(drop=True)


def detect_sudden_changes(
    df_ind: pd.DataFrame, target_date: pd.Timestamp | None = None,
    top_k: int = 20,
) -> pd.DataFrame:
    """Detección de cambios abruptos mensuales.

    Para cada cooperativa se calcula la diferencia respecto al mes anterior en
    cada indicador. Se toma como score la mayor variación absoluta observada
    entre todos los indicadores y se retorna el ranking descendente.
    """
    df = _filter_active(df_ind).dropna(subset=["date"])
    if df.empty:
        return pd.DataFrame()

    if target_date is None:
        target_date = df["date"].max()

    # Mes anterior
    prev_date = (target_date - pd.DateOffset(months=1)).normalize().replace(day=1)

    df_t = df[df["date"] == target_date].drop_duplicates(subset=["coop"], keep="first")
    df_p = df[df["date"] == prev_date].drop_duplicates(subset=["coop"], keep="first")
    if df_t.empty or df_p.empty:
        return pd.DataFrame()

    df_t = df_t.set_index("coop")[INDICATOR_KEYS]
    df_p = df_p.set_index("coop")[INDICATOR_KEYS]

    # Diferencias solo para cooperativas comunes
    common = df_t.index.intersection(df_p.index)
    if len(common) == 0:
        return pd.DataFrame()
    diff = df_t.loc[common] - df_p.loc[common]

    # Se descartan filas donde todos los indicadores sean NA
    diff = diff.dropna(how="all")
    if diff.empty:
        return pd.DataFrame()
    common = diff.index

    # Por cooperativa: mayor variación absoluta y su indicador (ignorando NA)
    abs_diff = diff.abs()
    max_change = abs_diff.max(axis=1, skipna=True)
    max_indicator = abs_diff.idxmax(axis=1, skipna=True)
    # Dirección real (con signo)
    signed_change = pd.Series(
        [diff.loc[c, max_indicator[c]] for c in common], index=common,
    )

    result = pd.DataFrame({
        "coop": common,
        "date": target_date,
        "max_abs_change": max_change.values,
        "max_signed_change": signed_change.values,
        "top_indicator": max_indicator.map(lambda k: INDICATORS[k]["short"]).values,
        "direction": ["Salto al alza" if v > 0 else "Caída abrupta" for v in signed_change.values],
    })

    # Se adjuntan las variaciones de cada indicador
    for k in INDICATOR_KEYS:
        result[f"diff_{k}"] = diff[k].values

    # Se excluyen finalmente las filas con max_change NA
    result = result.dropna(subset=["max_abs_change"])
    return result.sort_values("max_abs_change", ascending=False).head(top_k).reset_index(drop=True)


# =========================================================
# 10. F-04 LiNGAM (solo Shapiro-Wilk)
# =========================================================
def shapiro_wilk_table(df_ind: pd.DataFrame, target_date: pd.Timestamp | None = None) -> pd.DataFrame:
    """Test de Shapiro-Wilk para cada indicador (sin curtosis ni asimetría)."""
    df = _filter_active(df_ind).dropna(subset=["date"])
    if df.empty:
        return pd.DataFrame()

    if target_date is None:
        target_date = df["date"].max()

    df_t = df[df["date"] == target_date].drop_duplicates(subset=["coop"], keep="first")

    rows = []
    for k in INDICATOR_KEYS:
        x = df_t[k].dropna().values
        n = len(x)
        if n < 20:
            rows.append({"Indicador": INDICATORS[k]["short"], "n": n,
                        "p-valor Shapiro-Wilk": None, "No Gausiano": None})
            continue
        try:
            _, p = stats.shapiro(x)
            rows.append({
                "Indicador": INDICATORS[k]["short"], "n": n,
                "p-valor Shapiro-Wilk": round(float(p), 4),
                "No Gausiano": "✓" if p < 0.05 else "—",
            })
        except Exception:
            rows.append({"Indicador": INDICATORS[k]["short"], "n": n,
                        "p-valor Shapiro-Wilk": None, "No Gausiano": None})
    return pd.DataFrame(rows)


def run_lingam(
    df_ind: pd.DataFrame, target_date: pd.Timestamp | None = None,
) -> tuple[np.ndarray | None, list[str], pd.DataFrame | None]:
    """Ejecuta DirectLiNGAM y retorna la matriz de adyacencia y la lista de indicadores."""
    if not LINGAM_OK:
        return None, [], None

    df = _filter_active(df_ind).dropna(subset=["date"])
    if df.empty:
        return None, [], None

    if target_date is None:
        target_date = df["date"].max()

    df_t = df[df["date"] == target_date].drop_duplicates(subset=["coop"], keep="first")
    X = df_t[INDICATOR_KEYS].copy()
    # LiNGAM no admite faltantes → se imputan con la mediana
    X = X.fillna(X.median(numeric_only=True))
    X = X.dropna()  # Si aún quedan faltantes, se descartan
    if len(X) < 10:
        return None, [], None

    scaler = StandardScaler()
    X_scaled = scaler.fit_transform(X)

    try:
        model = lingam.DirectLiNGAM()
        model.fit(X_scaled)
        adj = model.adjacency_matrix_
        labels = [INDICATORS[k]["short"] for k in INDICATOR_KEYS]
        return adj, labels, X
    except Exception as e:
        st.warning(f"Error en cálculo de LiNGAM: {e}")
        return None, [], None


def lingam_to_plotly(
    adj: np.ndarray, labels: list[str], threshold: float = 0.15,
) -> go.Figure:
    """Representa la matriz de adyacencia de LiNGAM como red en Plotly."""
    n = len(labels)
    # Disposición circular
    angles = np.linspace(0, 2 * np.pi, n, endpoint=False)
    x = np.cos(angles)
    y = np.sin(angles)

    edge_x = []
    edge_y = []
    edge_text = []
    arrow_annotations = []
    for i in range(n):
        for j in range(n):
            coef = adj[i, j]
            if i == j or abs(coef) < threshold:
                continue
            # j → i (convención LiNGAM: adj[i,j] es el efecto de j sobre i)
            x0, y0 = x[j], y[j]
            x1, y1 = x[i], y[i]
            edge_x.extend([x0, x1, None])
            edge_y.extend([y0, y1, None])
            edge_text.append(f"{labels[j]} → {labels[i]}: {coef:.3f}")
            # Flecha
            color = "#C62828" if coef > 0 else "#1565C0"
            arrow_annotations.append(dict(
                x=x1 * 0.85, y=y1 * 0.85, ax=x0 * 0.85, ay=y0 * 0.85,
                xref="x", yref="y", axref="x", ayref="y",
                showarrow=True, arrowhead=3, arrowsize=1.5,
                arrowwidth=max(1.5, abs(coef) * 5),
                arrowcolor=color, opacity=0.7,
            ))

    fig = go.Figure()
    fig.add_trace(go.Scatter(
        x=x, y=y, mode="markers+text",
        marker=dict(size=40, color="#FFD966", line=dict(color="#404040", width=2)),
        text=labels, textposition="middle center",
        textfont=dict(size=12, color="black"),
        hoverinfo="text", showlegend=False,
    ))
    fig.update_layout(
        title=f"Grafo causal LiNGAM (|coef| > {threshold:.2f})",
        showlegend=False,
        annotations=arrow_annotations,
        xaxis=dict(showgrid=False, zeroline=False, showticklabels=False, range=[-1.5, 1.5]),
        yaxis=dict(showgrid=False, zeroline=False, showticklabels=False, range=[-1.5, 1.5],
                  scaleanchor="x", scaleratio=1),
        height=480, margin=dict(l=20, r=20, t=50, b=20),
        plot_bgcolor="white",
    )
    return fig


def lingam_to_graphviz(
    adj: np.ndarray, labels: list[str], threshold: float = 0.15,
) -> "graphviz.Digraph | None":
    """Dibuja la matriz de adyacencia de LiNGAM como DAG con Graphviz.

    - Coeficiente positivo: flecha azul (#1565C0)
    - Coeficiente negativo: flecha roja (#C62828)
    - A mayor |coeficiente|, más grueso el trazo; la etiqueta muestra el coeficiente.

    Exclusivo para la pestaña F-04.
    """
    if not GRAPHVIZ_OK:
        return None
    n = len(labels)
    dot = graphviz.Digraph("LiNGAM_DAG", format="svg")
    dot.attr(
        rankdir="LR",
        bgcolor="white",
        fontname="Helvetica",
        labelloc="t",
        label=f"Grafo causal LiNGAM (|coef| ≥ {threshold:.2f})",
        fontsize="14",
    )
    dot.attr(
        "node",
        shape="ellipse",
        style="filled",
        fillcolor="#FFD966",
        color="#404040",
        fontname="Helvetica",
        fontsize="11",
    )
    dot.attr("edge", fontname="Helvetica", fontsize="10")

    # Nodos
    for lbl in labels:
        dot.node(str(lbl), str(lbl))

    # Aristas (j → i: adj[i,j] es el efecto de j sobre i)
    for i in range(n):
        for j in range(n):
            coef = float(adj[i, j])
            if i == j or abs(coef) < threshold:
                continue
            color = "#1565C0" if coef > 0 else "#C62828"  # Positivo=azul, Negativo=rojo
            penwidth = f"{max(1.0, abs(coef) * 4):.2f}"
            dot.edge(
                str(labels[j]), str(labels[i]),
                label=f" {coef:+.3f} ",
                color=color,
                fontcolor=color,
                penwidth=penwidth,
                arrowsize="0.9",
            )
    return dot


# =========================================================
# 10b. LiNGAM por agrupación anual / diagnóstico de no gausianidad
# =========================================================
# Contexto:
#   Un LiNGAM ejecutado sobre una única instantánea mensual (target_date)
#   presenta dos problemas:
#   (a) Ruido mensual (cierres contables / estacionalidad) que hace inestables
#       los resultados.
#   (b) Muestra pequeña (una sola instantánea).
#
# Método:
#   Se agrupan todos los registros de cooperativa × mes del año objetivo en
#   una sola matriz de diseño y se ejecuta DirectLiNGAM. Al aparecer la misma
#   cooperativa varios meses, teóricamente la independencia se viola, pero se
#   prioriza el aseguramiento de la muestra (hasta 12× el número de cooperativas)
#   y la captura de cambios estructurales anuales. Es un método pragmático.
#
# Funciones públicas:
#   - run_lingam_year(df_ind, year)
#   - run_lingam_raw_year(df_raw_all_clean, selected_vars, year)
#   - shapiro_wilk_table_year(df_ind, year)
#   - available_years_from_ind(df_ind) / available_years_from_raw(df_raw)
#   - default_lingam_year(available_years)

DEFAULT_LINGAM_YEAR = 2025


def available_years_from_ind(df_ind: pd.DataFrame) -> list[int]:
    """Lista ascendente de años presentes en los datos de indicadores."""
    if df_ind.empty or "date" not in df_ind.columns:
        return []
    d = pd.to_datetime(df_ind["date"], errors="coerce").dropna()
    if d.empty:
        return []
    return sorted({int(x.year) for x in d})


def available_years_from_raw(df_raw_all_clean: pd.DataFrame) -> list[int]:
    """Lista ascendente de años presentes en los datos brutos."""
    if df_raw_all_clean.empty or "date" not in df_raw_all_clean.columns:
        return []
    d = pd.to_datetime(df_raw_all_clean["date"], errors="coerce").dropna()
    if d.empty:
        return []
    return sorted({int(x.year) for x in d})


def default_lingam_year(years: list[int], preferred: int = DEFAULT_LINGAM_YEAR) -> int | None:
    """Determina el año predeterminado: preferred si está disponible,
    de lo contrario el año más reciente.
    """
    if not years:
        return None
    if preferred in years:
        return preferred
    return max(years)


def run_lingam_year(
    df_ind: pd.DataFrame, year: int,
) -> tuple[np.ndarray | None, list[str], pd.DataFrame | None, dict]:
    """Ejecuta LiNGAM a nivel de indicadores mediante el método de agrupación anual.

    Agrupa los datos de todos los meses (enero a diciembre) del año objetivo
    en un solo conjunto de cooperativa × mes y ejecuta DirectLiNGAM.

    Returns:
        (adj, labels, X_used, meta)
        - adj: Matriz de adyacencia (shape=(n_ind, n_ind)) o None
        - labels: Nombres de indicadores
        - X_used: Matriz efectivamente usada (DataFrame)
        - meta: {"year", "n_samples", "n_months", "months", "n_coops"}
    """
    meta = {"year": int(year), "n_samples": 0, "n_months": 0, "months": [], "n_coops": 0}
    if not LINGAM_OK:
        return None, [], None, meta

    df = _filter_active(df_ind).dropna(subset=["date"]).copy()
    if df.empty:
        return None, [], None, meta

    df["date"] = pd.to_datetime(df["date"], errors="coerce")
    df = df.dropna(subset=["date"])
    df_y = df[df["date"].dt.year == int(year)]
    if df_y.empty:
        return None, [], None, meta

    # Si hay duplicados (coop, date), se toma la primera ocurrencia
    df_y = df_y.drop_duplicates(subset=["coop", "date"], keep="first")
    X = df_y[INDICATOR_KEYS].copy()
    # LiNGAM no admite faltantes → se imputan con la mediana
    X = X.fillna(X.median(numeric_only=True))
    X = X.dropna()
    if len(X) < 10:
        return None, [], None, meta

    months_list = sorted({pd.Timestamp(d).strftime("%Y-%m") for d in df_y["date"]})
    n_coops = int(df_y["coop"].nunique())

    scaler = StandardScaler()
    X_scaled = scaler.fit_transform(X)

    try:
        model = lingam.DirectLiNGAM()
        model.fit(X_scaled)
        adj = model.adjacency_matrix_
        labels = [INDICATORS[k]["short"] for k in INDICATOR_KEYS]
        meta.update({
            "n_samples": int(len(X)),
            "n_months": len(months_list),
            "months": months_list,
            "n_coops": n_coops,
        })
        return adj, labels, X, meta
    except Exception as e:
        st.warning(f"Error en cálculo de LiNGAM: {e}")
        return None, [], None, meta


def run_lingam_raw_year(
    df_raw_all_clean: pd.DataFrame,
    selected_vars: list[str],
    year: int,
) -> tuple[np.ndarray | None, list[str], pd.DataFrame | None, dict]:
    """Ejecuta LiNGAM a nivel de variables brutas mediante el método de agrupación anual.

    Agrupa los 12 meses del año objetivo en un conjunto cooperativa × mes y
    conserva solo las filas donde estén presentes todas las selected_vars,
    luego ejecuta DirectLiNGAM.
    """
    meta = {"year": int(year), "n_samples": 0, "n_months": 0, "months": [], "n_coops": 0}
    if not LINGAM_OK or len(selected_vars) < 2 or df_raw_all_clean.empty:
        return None, [], None, meta

    df = df_raw_all_clean[~df_raw_all_clean["is_liquidating"]].copy()
    if df.empty or "date" not in df.columns:
        return None, [], None, meta

    df["date"] = pd.to_datetime(df["date"], errors="coerce")
    df = df.dropna(subset=["date"])
    df_y = df[df["date"].dt.year == int(year)]
    df_y = df_y[df_y["variable"].isin(selected_vars)]
    if df_y.empty:
        return None, [], None, meta

    # Pivot por cooperativa × mes: filas=(coop,date), columnas=variable
    wide = df_y.pivot_table(
        index=["coop", "date"], columns="variable", values="value", aggfunc="first",
    )
    avail = [v for v in selected_vars if v in wide.columns]
    if len(avail) < 2:
        return None, avail, None, meta
    wide = wide[avail].dropna()
    if len(wide) < 10:
        return None, avail, None, meta

    # Meta: meses y cooperativas incluidas
    idx = wide.index.to_frame(index=False)
    months_list = sorted({pd.Timestamp(d).strftime("%Y-%m") for d in idx["date"]})
    n_coops = int(idx["coop"].nunique())

    scaler = StandardScaler()
    X_scaled = scaler.fit_transform(wide.values)
    try:
        model = lingam.DirectLiNGAM()
        model.fit(X_scaled)
        adj = model.adjacency_matrix_
        meta.update({
            "n_samples": int(len(wide)),
            "n_months": len(months_list),
            "months": months_list,
            "n_coops": n_coops,
        })
        return adj, avail, wide, meta
    except Exception as e:
        st.warning(f"Error en cálculo de LiNGAM: {e}")
        return None, avail, None, meta


def shapiro_wilk_table_year(df_ind: pd.DataFrame, year: int) -> pd.DataFrame:
    """Aplica Shapiro-Wilk a cada indicador sobre la agrupación anual (coop × mes)."""
    df = _filter_active(df_ind).dropna(subset=["date"]).copy()
    if df.empty:
        return pd.DataFrame()
    df["date"] = pd.to_datetime(df["date"], errors="coerce")
    df = df.dropna(subset=["date"])
    df_y = df[df["date"].dt.year == int(year)]
    if df_y.empty:
        return pd.DataFrame()
    df_y = df_y.drop_duplicates(subset=["coop", "date"], keep="first")

    rows = []
    for k in INDICATOR_KEYS:
        x = df_y[k].dropna().values
        n = len(x)
        if n < 20:
            rows.append({"Indicador": INDICATORS[k]["short"], "n": n,
                         "p-valor Shapiro-Wilk": None, "No Gausiano": None})
            continue
        try:
            _, p = stats.shapiro(x)
            rows.append({
                "Indicador": INDICATORS[k]["short"], "n": n,
                "p-valor Shapiro-Wilk": round(float(p), 4),
                "No Gausiano": "✓" if p < 0.05 else "—",
            })
        except Exception:
            rows.append({"Indicador": INDICATORS[k]["short"], "n": n,
                         "p-valor Shapiro-Wilk": None, "No Gausiano": None})
    return pd.DataFrame(rows)


# =========================================================
# 11. Agregación de datos para el Tablero
# =========================================================
def build_dashboard_data(df_ind: pd.DataFrame) -> dict:
    """Retorna los datos agregados para el Tablero."""
    if df_ind.empty:
        return {}

    df = _filter_active(df_ind).dropna(subset=["date"])
    if df.empty:
        return {}

    latest = df["date"].max()
    # Variación interanual (YoY): permite observar los cambios estructurales
    # del sector eliminando estacionalidad, en vez de la variación mensual.
    yoy = (latest - pd.DateOffset(years=1)).normalize().replace(day=1)

    # Panel A: mediana de los 5 indicadores + flecha de variación interanual
    pane_a = []
    df_l = df[df["date"] == latest]
    df_yoy = df[df["date"] == yoy]
    for k, cfg in INDICATORS.items():
        val_l = df_l[k].median(skipna=True)
        val_yoy = df_yoy[k].median(skipna=True) if not df_yoy.empty else np.nan
        arrow = _trend_arrow(val_l, val_yoy)
        pane_a.append({
            "indicator": cfg["short"],
            "key": k,
            "value": val_l,
            "value_yoy": val_yoy,
            "arrow": arrow,
            "percent": cfg["percent"],
            "benchmark": cfg["benchmark"],
            "direction": cfg["direction"],
        })

    n_active = df_l["coop"].nunique()
    n_liq = df_ind[(df_ind["date"] == latest) & df_ind["is_liquidating"]]["coop"].nunique()

    # Panel C: TOP 10 Anomalías
    anom_iso = detect_anomalies_isoforest(df_ind, latest)
    top10_anom = anom_iso.head(10) if not anom_iso.empty else pd.DataFrame()

    # Número de indicadores no gausianos (basado en el año objetivo del LiNGAM)
    years_ind = available_years_from_ind(df_ind)
    lingam_year = default_lingam_year(years_ind, preferred=DEFAULT_LINGAM_YEAR)
    if lingam_year is not None:
        sw_table = shapiro_wilk_table_year(df_ind, lingam_year)
    else:
        sw_table = pd.DataFrame()
    if not sw_table.empty:
        valid = sw_table.dropna(subset=["No Gausiano"])
        n_total = len(valid)
        n_ng = (valid["No Gausiano"] == "✓").sum()
    else:
        n_total = n_ng = 0

    return {
        "latest_date": latest,
        "yoy_date": yoy,
        "n_active": n_active,
        "n_liq": n_liq,
        "pane_a": pane_a,
        "top10_anom": top10_anom,
        "n_nongaussian": int(n_ng),
        "n_total_ind": int(n_total),
        "lingam_year": lingam_year,
    }


# =========================================================
# 12. Interfaz principal
# =========================================================
st.title("🏦 Plataforma de Análisis de Estados Financieros")
st.caption(
    "Tablero como pantalla principal / 4 módulos (F-01 a F-04) / "
    f"LLM = `{LLM_MODEL}` local / "
    "LiNGAM = método de agrupación anual (predeterminado: 2025)"
)


# ---- Configuración fija (sin barra lateral) ----
# Umbrales de unificación de nombres (valores provisionales tras verificación)
threshold_auto = 85
threshold_review = 70
# Configuración LLM (valores locales predeterminados)
llm_host = ""       # Vacío → http://localhost:11434
llm_timeout = 180   # Segundos


# ---- Carga de archivo ----
uploaded_file = st.file_uploader(
    "📥 Arrastre y suelte el archivo Excel de Estados Financieros",
    type=["xlsx"],
    help="Estados financieros mensuales en formato publicado por CONSUCOOP",
)

if uploaded_file is None:
    st.info(
        "⬆️ Al cargar el archivo Excel, se ejecutará automáticamente la "
        "unificación de nombres y el cálculo de indicadores financieros.\n\n"
        "Formatos de hoja admitidos:\n"
        "- `MM-YYYY` (2019-2023) … BG + ER en una sola hoja\n"
        "- `BG MM-YYYY` / `ER MM-YYYY` (2024 en adelante) … BG y ER en hojas separadas"
    )
    st.stop()


# ---- Lectura de datos ----
with st.spinner("📂 Leyendo archivo Excel..."):
    names, sheet_info, df_raw, df_raw_all = load_excel(uploaded_file.read())

if not names:
    st.error("❌ No se pudieron cargar nombres de cooperativas. Verifique el formato de las hojas.")
    st.stop()


# ---- Unificación de nombres ----
with st.spinner("🔗 Unificando nombres..."):
    clusters, review = cluster_names(
        names, threshold_auto=threshold_auto, threshold_review=threshold_review,
    )
    name_mapping = build_name_mapping(clusters)


# ---- Cálculo de indicadores ----
with st.spinner("🧮 Calculando indicadores financieros..."):
    df_ind = compute_indicators(df_raw, name_mapping)

# ---- Procesamiento del formato largo de variables brutas ----
with st.spinner("🧾 Procesando variables brutas..."):
    df_raw_all_clean = compute_raw_all(df_raw_all, name_mapping)


# ---- Indicadores clave (KPI) superiores ----
st.markdown("### 📊 Resumen")
col1, col2, col3, col4, col5, col6 = st.columns(6)
col1.metric("Nombres (brutos)", f"{len(names):,}")
col2.metric("Clústeres tras fusión", f"{len(clusters):,}",
            delta=f"{len(clusters) - len(names):+,}", delta_color="inverse")
col3.metric("Pares por revisar", f"{len(review):,}")
liq_count = sum(1 for c in clusters if c["is_liquidating"])
col4.metric("Clústeres en liquidación", f"{liq_count:,}")
if not df_ind.empty:
    col5.metric("Registros de indicadores", f"{len(df_ind):,}")
    col6.metric("Mes más reciente",
                f"{df_ind['date'].max():%Y-%m}" if df_ind["date"].notna().any() else "N/A")
else:
    col5.metric("Registros de indicadores", "0")
    col6.metric("Mes más reciente", "N/A")


# ---- Composición de pestañas (Tablero + 5 pestañas) ----
tab_dashboard, tab_data_quality, tab_indicators, tab_timeseries, tab_anomaly, tab_lingam = st.tabs([
    "🏠 Tablero",
    "📂 Calidad de Datos",
    "📊 F-01 Indicadores",
    "📈 F-02 Series de Tiempo",
    "⚠️ F-03 Detección de Anomalías",
    "🕸️ F-04 LiNGAM",
])


# =========================================================
# 13.1 🏠 Tablero (pantalla principal)
# =========================================================
with tab_dashboard:
    if df_ind.empty:
        st.error("No hay datos suficientes para calcular los indicadores financieros.")
    else:
        with st.spinner("Generando Tablero..."):
            dash = build_dashboard_data(df_ind)

        if not dash:
            st.warning("No hay datos válidos disponibles.")
        else:
            # Encabezado
            header_cols = st.columns([3, 2, 2, 2])
            header_cols[0].markdown(f"#### 📅 Mes más reciente: **{dash['latest_date']:%B %Y}**")
            header_cols[1].metric("Cooperativas activas", f"{dash['n_active']:,}")
            header_cols[2].metric(
                "En liquidación", f"{dash['n_liq']:,}",
                help="Excluidas de mediana, anomalías y LiNGAM",
            )
            header_cols[3].metric("Registros de indicadores", f"{len(df_ind):,}")
            st.divider()

            # Tablero con 2 paneles (el detalle de LiNGAM se encuentra en F-04)
            pane_a_col, pane_c_col = st.columns([1, 1])

            # === Panel A: Situación financiera central ===
            with pane_a_col:
                st.markdown("##### 📊 A. Tendencia general de situaciones financieras")
                st.caption(
                    f"Mediana de los 5 indicadores + flecha de variación interanual "
                    f"(objetivo: {dash['n_active']} cooperativas; "
                    f"{dash['n_liq']} en liquidación excluidas / "
                    f"comparación con {dash['yoy_date']:%Y-%m})"
                )
                for item in dash["pane_a"]:
                    val_str = _fmt_value(item["value"], item["percent"])
                    arrow = item["arrow"]
                    # Fórmula del indicador (descripción breve para la UI)
                    formula = INDICATORS[item["key"]]["formula"]
                    # Comparación con benchmark
                    if not (np.isnan(item["value"]) if isinstance(item["value"], float) else item["value"] is None):
                        if item["direction"] == "higher_better":
                            ok = item["value"] >= item["benchmark"]
                        else:
                            ok = item["value"] <= item["benchmark"]
                        bg = "#E8F5E9" if ok else "#FFEBEE"
                    else:
                        bg = "#F5F5F5"
                    st.markdown(
                        f"<div style='background-color:{bg};padding:8px 12px;margin:4px 0;border-radius:4px;'>"
                        f"<b>{item['indicator']}</b> &nbsp;&nbsp; {val_str} &nbsp; {arrow}"
                        f"<div style='font-size:0.80em;color:#555;margin-top:2px;'>{formula}</div>"
                        f"</div>",
                        unsafe_allow_html=True,
                    )

            # === Panel C: TOP 10 Anomalías ===
            with pane_c_col:
                st.markdown("##### ⚠️ B. 10 Anomalías en el mes más reciente en base al aprendizaje automático")
                if dash["top10_anom"].empty:
                    st.info("Datos insuficientes para detección de anomalías.")
                else:
                    df_disp = dash["top10_anom"].head(10).copy()
                    df_disp["Posición"] = range(1, len(df_disp) + 1)
                    df_disp["Score"] = df_disp["anomaly_score"].round(3)
                    df_disp["Cooperativa"] = df_disp["coop"].str.slice(0, 28) + "..."
                    df_disp["Posible irregularidad"] = df_disp["top_concern"]
                    st.dataframe(
                        df_disp[["Posición", "Cooperativa", "Score", "Posible irregularidad"]],
                        use_container_width=True, hide_index=True, height=380,
                    )


# =========================================================
# 13.2 📂 F-00 Calidad de Datos
# =========================================================
with tab_data_quality:
    sub_clusters, sub_review, sub_single, sub_sheets, sub_mapping = st.tabs([
        "🔗 Clústeres Fusionados", "🔎 Pares por Revisar", "📎 Clústeres Individuales",
        "📋 Lista de Hojas", "💾 CSV de Mapeo",
    ])

    with sub_clusters:
        st.subheader("Clústeres con fusión")
        multi = [c for c in clusters if c["member_count"] >= 2]
        if not multi:
            st.info("No se produjeron fusiones en ningún clúster.")
        else:
            multi_sorted = sorted(multi, key=lambda x: -x["member_count"])
            df_multi = pd.DataFrame([{
                "Nº Miembros": c["member_count"],
                "Nombre Representativo": c["representative"],
                "En Liquidación": "✓" if c["is_liquidating"]
                          else ("⚠ parcial" if c["has_liquidating_variant"] else ""),
                "Variantes Fusionadas": " / ".join(c["members"]),
            } for c in multi_sorted])
            st.dataframe(df_multi, use_container_width=True, hide_index=True)

    with sub_review:
        st.subheader("Pares por revisar")
        st.caption(f"Rango de similitud {threshold_review}% a {threshold_auto - 1}%. Se muestran los 20 primeros.")
        if not review:
            st.success("✅ No hay pares que requieran revisión.")
        else:
            df_review = pd.DataFrame(
                review, columns=["Nombre de Cooperativa", "Coincidencia con", "Similitud (%)"]
            ).sort_values("Similitud (%)", ascending=False).head(20)
            df_review["Similitud (%)"] = df_review["Similitud (%)"].round(1)
            st.dataframe(df_review, use_container_width=True, hide_index=True)

    with sub_single:
        st.subheader("Clústeres individuales")
        single = [c for c in clusters if c["member_count"] == 1]
        df_single = pd.DataFrame({
            "Nombre Representativo": [c["representative"] for c in single],
            "En Liquidación": ["✓" if c["is_liquidating"] else "" for c in single],
        }).sort_values("Nombre Representativo")
        st.dataframe(df_single, use_container_width=True, hide_index=True)

    with sub_sheets:
        st.subheader("Hojas detectadas")
        df_sheets = pd.DataFrame(sheet_info)
        st.dataframe(df_sheets, use_container_width=True, hide_index=True)
        st.markdown("##### Resumen por formato")
        fmt_counts = df_sheets["format"].value_counts().reset_index()
        fmt_counts.columns = ["Formato", "Nº de Hojas"]
        st.dataframe(fmt_counts, use_container_width=True, hide_index=True)

    with sub_mapping:
        st.subheader("CSV de Mapeo")
        st.caption("Correspondencia nombre original → nombre representativo. En caso de fusión errónea, editar manualmente y recargar.")
        mapping_rows = []
        for c in clusters:
            for m in c["members"]:
                _, is_liq = extract_liquidating(m)
                mapping_rows.append({
                    "original": m,
                    "representative": c["representative"],
                    "cluster_id": c["cluster_id"],
                    "member_count": c["member_count"],
                    "is_liquidating_variant": is_liq,
                })
        df_mapping = pd.DataFrame(mapping_rows).sort_values(["representative", "original"])
        st.dataframe(df_mapping, use_container_width=True, hide_index=True)
        st.download_button(
            "📥 Descargar CSV de Mapeo",
            data=df_mapping.to_csv(index=False).encode("utf-8-sig"),
            file_name=f"coop_mapping_{datetime.now():%Y%m%d_%H%M}.csv",
            mime="text/csv",
            type="primary",
        )


# =========================================================
# 13.3 📊 F-01 Análisis de Indicadores Financieros (basado en mediana, Top 10)
# =========================================================
with tab_indicators:
    if df_ind.empty:
        st.error("No hay datos suficientes para calcular los indicadores financieros.")
    else:
        st.markdown(
            f"#### Datos objetivo: {df_ind['coop'].nunique():,} cooperativas × "
            f"{df_ind['date'].nunique():,} meses"
        )
        excl = df_ind[df_ind["is_liquidating"]]["coop"].nunique()
        st.caption(f"📌 {excl} cooperativas en liquidación excluidas de la mediana")
        st.caption(
            "📌 El valor representativo es la mediana. "
            "Para profundizar en cooperativas individuales, utilice F-02 (series de tiempo) "
            "o F-03 (detección de anomalías)."
        )

        ind_tabs = st.tabs([INDICATORS[k]["label"] for k in INDICATORS])
        for (key, cfg), t in zip(INDICATORS.items(), ind_tabs):
            with t:
                st.markdown(f"**Fórmula:** {cfg['formula']}")
                st.markdown(f"**Interpretación:** {cfg['explain']}")

                fig_med, df_med = build_sector_median_chart(df_ind, key)
                if not df_med.empty:
                    st.plotly_chart(fig_med, use_container_width=True, key=f"f01_med_{key}")

                    col_latest, col_mom = st.columns(2)
                    latest_row = df_med.sort_values("date").iloc[-1]
                    col_latest.metric(
                        f"{cfg['short']} del mes más reciente (mediana)",
                        _fmt_value(latest_row[key], cfg["percent"]),
                        help=f"{latest_row['date']:%Y-%m}",
                    )
                    if len(df_med) >= 2:
                        prev_row = df_med.sort_values("date").iloc[-2]
                        delta_mom = latest_row[key] - prev_row[key]
                        col_mom.metric(
                            "Variación mensual", _fmt_value(delta_mom, cfg["percent"]),
                            help=f"Comparación con {prev_row['date']:%Y-%m}",
                        )
                    else:
                        col_mom.metric("Variación mensual", "N/A")
                else:
                    st.warning("No hay datos válidos para este indicador.")

                with st.expander("Datos detallados (todas las cooperativas × meses)"):
                    detail = df_ind[["coop", "date", key, "is_liquidating"]].dropna(subset=[key])
                    detail = detail.sort_values(["date", "coop"])
                    detail[key] = detail[key].round(4)
                    st.dataframe(detail, use_container_width=True, hide_index=True)


# =========================================================
# 13.4 📈 F-02 Análisis de Series de Tiempo (MM 3/12, YoY)
# =========================================================
with tab_timeseries:
    if df_ind.empty:
        st.error("No hay datos suficientes para calcular los indicadores financieros.")
    else:
        st.markdown("#### 📈 Análisis de Series de Tiempo")
        st.caption(
            "Máximo 3 cooperativas y 1 indicador (para mantener la comparación legible) / "
            "Media móvil: opciones de 3 o 12 meses"
        )

        all_coops = sorted(_filter_active(df_ind)["coop"].unique())

        # Cooperativa predeterminada: ELGA (Cooperativa de Ahorro y Crédito Elga, Limitada)
        default_coop = next(
            (c for c in all_coops if "Elga" in c or "ELGA" in c.upper()),
            all_coops[0] if all_coops else None,
        )
        default_coops = [default_coop] if default_coop else []

        col_sel1, col_sel2 = st.columns([2, 1])
        with col_sel1:
            sel_coops = st.multiselect(
                "Seleccione cooperativas (máximo 3)", all_coops, max_selections=3,
                default=default_coops,
            )
        with col_sel2:
            # 1 solo indicador; predeterminado: Rentabilidad (ROA)
            sel_ind = st.radio(
                "Seleccione indicador (1)",
                INDICATOR_KEYS,
                index=INDICATOR_KEYS.index("roa"),
                format_func=lambda k: INDICATORS[k]["short"],
            )
            sel_inds = [sel_ind]

        col_opt1, col_opt2, col_opt3 = st.columns(3)
        with col_opt1:
            ma_choice = st.radio(
                "Media móvil", ["3 meses", "12 meses"], index=0, horizontal=True,
                help="Tendencia corta (3) o anual (12)",
            )
            ma_window = 3 if ma_choice == "3 meses" else 12
        with col_opt2:
            show_yoy = st.checkbox("Superponer YoY (%)", value=False)
        with col_opt3:
            zero_as_nan = st.checkbox("Tratar 0 como faltante", value=True)

        if not sel_coops:
            st.info("Seleccione una cooperativa (por defecto ya se selecciona ELGA).")
        else:
            fig_ts = build_time_series_chart(
                df_ind, sel_coops, sel_inds, ma_window, show_yoy, zero_as_nan,
            )
            st.plotly_chart(fig_ts, use_container_width=True, key="f02_ts")
            with st.expander("Lista de datos seleccionados"):
                df_show = df_ind[df_ind["coop"].isin(sel_coops)][
                    ["coop", "date"] + sel_inds
                ].dropna(subset=sel_inds, how="all").sort_values(["coop", "date"])
                st.dataframe(df_show, use_container_width=True, hide_index=True)


# =========================================================
# 13.5 ⚠️ F-03 Detección de Anomalías (Isolation Forest + cambios abruptos)
# =========================================================
with tab_anomaly:
    if df_ind.empty:
        st.error("No hay datos suficientes para calcular los indicadores financieros.")
    else:
        st.markdown("#### ⚠️ Detección de Anomalías (3 métodos)")

        with st.expander("📖 ¿Qué hace cada uno de los 3 métodos? (clic para expandir)", expanded=False):
            st.markdown("""
**① Isolation Forest (anomalía global de corte transversal a nivel de indicadores)**
- **Qué observa**: En un solo mes objetivo, posiciona a cada cooperativa en el espacio de los 5 indicadores (Liquidez, Solvencia, Rentabilidad, Crecimiento, Calidad de Cartera (Morosidad)) y detecta las que se alejan del conjunto.
- **No considera el cambio respecto al mes anterior** → es menos reactivo a picos puntuales por errores de registro.
- **Tipo de anomalía detectada**: Cooperativas que se apartan persistentemente del sector (tamaño muy pequeño, estructura muy sesgada, etc.).

**② Cambio abrupto a nivel de indicadores (corte longitudinal)**
- **Qué observa**: Detecta las cooperativas con mayor diferencia entre el mes objetivo y el mes anterior en cada **indicador**.
- **Capta picos puntuales** → se contaminan más con errores de registro y valores atípicos temporales.
- **Tipo de anomalía detectada**: Cooperativas que venían normales y tuvieron un salto brusco en el último mes.

**③ Variación mensual % de variables brutas (cambios abruptos a nivel de variable)**
- **Qué observa**: Top 20 de (cooperativa × variable bruta) con mayor valor absoluto de variación % mensual (=(mes actual − mes anterior) / |mes anterior|).
- **Usa % en vez de valor absoluto** → permite comparar variaciones amplias sin depender de la escala (millones vs miles).
- **Visualiza movimientos ocultos en los indicadores** al nivel original de cuenta contable.

**Cómo usar cada uno**:
- ① identifica posibles **cooperativas estructuralmente atípicas**, ② las que **tuvieron algo inusual el mes pasado**.
- ③ identifica cooperativas con **cambios abruptos en una cuenta específica**.

**⚠️ Cuando detecte una anomalía**: consulte la tabla de datos brutos de los últimos 3 meses que aparece más abajo.
"Salto brusco de orden de magnitud → posible error de registro".
"Deterioro asintótico → probable problema real de gestión".
            """.strip())

        df_active = _filter_active(df_ind).dropna(subset=["date"])
        latest_date = df_active["date"].max()

        col_param1, col_param2 = st.columns(2)
        with col_param1:
            target_str = st.selectbox(
                "Mes objetivo",
                options=sorted(df_active["date"].dropna().unique(), reverse=True),
                index=0, format_func=lambda d: pd.Timestamp(d).strftime("%Y-%m"),
            )
            target_date = pd.Timestamp(target_str)
        with col_param2:
            contamination = st.slider(
                "Tasa de contaminación de Isolation Forest",
                0.05, 0.30, 0.15, 0.05,
                help="Proporción esperada de cooperativas anómalas",
            )

        sub_iso, sub_sudden, sub_raw_pct = st.tabs([
            "① Isolation Forest (indicadores, global)",
            "② Cambio abrupto mensual (indicadores)",
            "③ Variación mensual % de variables brutas",
        ])

        # === ① Isolation Forest ===
        with sub_iso:
            st.info(
                "💡 **Este método no observa el cambio respecto al mes anterior**. "
                "Detecta las cooperativas cuyo patrón de 5 indicadores en el mes objetivo "
                "se aleja del resto."
            )
            with st.spinner("Calculando Isolation Forest..."):
                anom_iso = detect_anomalies_isoforest(
                    df_ind, target_date, contamination=contamination,
                )
            if anom_iso.empty:
                st.warning("Datos insuficientes para el análisis.")
            else:
                st.markdown(f"##### Ranking de anomalía global de {target_date:%Y-%m} (Top 20)")
                df_iso_disp = anom_iso.head(20).copy()
                df_iso_disp["Score de Anomalía"] = df_iso_disp["anomaly_score"].round(3)
                df_iso_disp["Posible irregularidad"] = df_iso_disp["top_concern"]
                df_iso_disp["Anomalía"] = df_iso_disp["is_anomaly"].map({True: "⚠️", False: ""})
                df_iso_disp.insert(0, "Posición", range(1, len(df_iso_disp) + 1))
                st.dataframe(
                    df_iso_disp[["Posición", "coop", "Score de Anomalía", "Posible irregularidad", "Anomalía"]]
                    .rename(columns={"coop": "Cooperativa"}),
                    use_container_width=True, hide_index=True,
                )

                # Contribución de la cooperativa seleccionada
                st.markdown("##### 🔍 Detalle de la Cooperativa Seleccionada")
                sel_coop = st.selectbox(
                    "Seleccione cooperativa", df_iso_disp["coop"].tolist(), key="iso_sel",
                )

                # Datos brutos de los últimos 3 meses + indicadores
                st.markdown("**📋 Datos brutos de los últimos 3 meses (principales) + Indicadores**")
                st.caption(
                    "⚠️ Si observa valores de orden de magnitud muy diferente o saltos bruscos, "
                    "puede tratarse de un error de registro en el Excel original. "
                    "Si el deterioro es asintótico, es probable que sea un problema real de gestión."
                )
                raw_table = get_recent_raw_data(df_ind, sel_coop, target_date, n_months=3)
                if not raw_table.empty:
                    st.dataframe(raw_table, use_container_width=True, hide_index=True)
                else:
                    st.info("No se encontraron datos recientes de esta cooperativa.")

                # Datos brutos completos (todas las cuentas)
                st.markdown("**📋 Datos brutos de los últimos 3 meses (todas las cuentas, miles HNL)**")
                st.caption(
                    "Los 7 rubros principales no siempre permiten identificar 'dónde ocurre la anomalía'. "
                    "Se muestran todas las cuentas de BG/ER. Use la columna de variación mensual % (derecha) "
                    "para encontrar los movimientos amplios."
                )
                all_raw_table = get_recent_all_raw_data(
                    df_raw_all_clean, sel_coop, target_date, n_months=3,
                )
                if not all_raw_table.empty:
                    st.dataframe(all_raw_table, use_container_width=True, hide_index=True, height=420)
                else:
                    st.info("No se encontraron datos brutos de esta cooperativa.")

                # Gráfico de contribución
                st.markdown("**📊 Contribución por indicador (valor absoluto estandarizado)**")
                row = anom_iso[anom_iso["coop"] == sel_coop].iloc[0]
                contrib_data = pd.DataFrame({
                    "Indicador": [INDICATORS[k]["short"] for k in INDICATOR_KEYS],
                    "Contribución": [row[f"contrib_{k}"] for k in INDICATOR_KEYS],
                })
                fig_c = go.Figure(go.Bar(
                    x=contrib_data["Contribución"], y=contrib_data["Indicador"],
                    orientation="h", marker=dict(color="#C62828"),
                    text=[f"{v:.2f}" for v in contrib_data["Contribución"]],
                    textposition="outside",
                ))
                fig_c.update_layout(
                    title=f"Contribución de {sel_coop[:50]}",
                    height=280, margin=dict(l=20, r=20, t=40, b=20),
                )
                st.plotly_chart(fig_c, use_container_width=True, key="f03_contrib")

                # CSV
                st.download_button(
                    "📥 Descargar CSV de Isolation Forest",
                    data=anom_iso.to_csv(index=False).encode("utf-8-sig"),
                    file_name=f"anomaly_isoforest_{target_date:%Y%m}.csv",
                    mime="text/csv",
                )

        # === ② Cambio abrupto mensual ===
        with sub_sudden:
            st.info(
                "💡 **Este método sí observa el cambio respecto al mes anterior**. "
                "Detecta las cooperativas con mayor diferencia mensual en cualquiera de los indicadores. "
                "Como los errores de registro se capturan con facilidad, revise siempre los datos brutos de abajo."
            )
            with st.spinner("Calculando cambios abruptos mensuales..."):
                sudden = detect_sudden_changes(df_ind, target_date, top_k=20)
            if sudden.empty:
                st.warning(
                    "No hay datos para comparar con el mes anterior. "
                    "Puede que ninguna cooperativa tenga datos en el mes previo al objetivo."
                )
            else:
                st.markdown(f"##### Ranking de cambios abruptos de {target_date:%Y-%m} (Top 20)")
                st.caption(
                    "Se calcula la diferencia mensual por cooperativa e indicador, y se presenta a quienes tienen "
                    "la mayor variación absoluta. 'Salto al alza' / 'Caída abrupta' indican solo la dirección; "
                    "la bondad o gravedad depende del indicador."
                )
                df_s = sudden.copy()
                df_s.insert(0, "Posición", range(1, len(df_s) + 1))
                df_s["Mayor Variación Absoluta"] = df_s["max_abs_change"].round(4)
                df_s["Dirección"] = df_s["direction"]
                df_s["Indicador Principal"] = df_s["top_indicator"]
                st.dataframe(
                    df_s[["Posición", "coop", "Indicador Principal", "Dirección", "Mayor Variación Absoluta"]]
                    .rename(columns={"coop": "Cooperativa"}),
                    use_container_width=True, hide_index=True,
                )

                # Datos de la cooperativa seleccionada (últimos 3 meses)
                st.markdown("##### 🔍 Detalle de la Cooperativa Seleccionada")
                sel_coop_s = st.selectbox(
                    "Seleccione cooperativa", df_s["coop"].tolist(), key="sudden_sel",
                )
                st.markdown("**📋 Datos brutos de los últimos 3 meses (principales) + Indicadores**")
                st.caption(
                    "⚠️ Los cambios abruptos mensuales suelen reflejar errores de registro. "
                    "Salto de un solo mes con orden de magnitud distinto → probable error de registro. "
                    "Deterioro sostenido → probable problema real de gestión."
                )
                raw_table_s = get_recent_raw_data(df_ind, sel_coop_s, target_date, n_months=3)
                if not raw_table_s.empty:
                    st.dataframe(raw_table_s, use_container_width=True, hide_index=True)
                else:
                    st.info("No se encontraron datos recientes de esta cooperativa.")

                # Todas las cuentas
                st.markdown("**📋 Datos brutos de los últimos 3 meses (todas las cuentas, miles HNL)**")
                st.caption(
                    "Para identificar en qué cuenta ocurrió el cambio, se muestran todas las cuentas de BG/ER."
                )
                all_raw_table_s = get_recent_all_raw_data(
                    df_raw_all_clean, sel_coop_s, target_date, n_months=3,
                )
                if not all_raw_table_s.empty:
                    st.dataframe(all_raw_table_s, use_container_width=True, hide_index=True, height=420)
                else:
                    st.info("No se encontraron datos brutos de esta cooperativa.")

                # CSV
                st.download_button(
                    "📥 Descargar CSV de Cambios Abruptos",
                    data=sudden.to_csv(index=False).encode("utf-8-sig"),
                    file_name=f"anomaly_sudden_{target_date:%Y%m}.csv",
                    mime="text/csv",
                )

        # === ③ Variación mensual % de variables brutas ===
        with sub_raw_pct:
            st.info(
                "💡 **Este método observa la variación mensual % de variables brutas** "
                "(= (mes actual − mes anterior) / |mes anterior|). "
                "Al comparar por % en lugar de valor absoluto, se pueden colocar en un mismo marco "
                "los 'movimientos amplios' aunque las variables tengan escalas distintas. Se muestran los 20 primeros."
            )
            st.caption(
                "⚠️ **Tratamiento de ceros / valores pequeños**: Si el mes actual o el anterior "
                "tienen valor 0 o un valor muy pequeño (|value| < 1 mil HNL = 1,000 HNL), la tasa de "
                "variación resulta extrema (o no calculable). Como muchas veces los faltantes se "
                "registran como 0, **se excluyen las filas donde cualquiera de los dos meses tenga "
                "valor 0, faltante, o |value| < 1 mil HNL** (valores que se visualizarían como '0' o '-0' "
                "al redondear). Solo se consideran filas con valores válidos y suficientemente grandes "
                "para producir el Top 20."
            )
            if df_raw_all_clean.empty:
                st.warning("No se cargaron datos brutos.")
            else:
                with st.spinner("Calculando variación mensual % de variables brutas..."):
                    raw_pct_df = detect_raw_pct_changes(
                        df_raw_all_clean, target_date, top_k=20,
                    )
                if raw_pct_df.empty:
                    st.warning(
                        "No hay registros comparables: o no existen datos del mes anterior, "
                        "o todos los valores del mes actual/anterior son cero o faltantes."
                    )
                else:
                    st.markdown(
                        f"##### Variables brutas con mayor variación mensual % en {target_date:%Y-%m} (Top 20)"
                    )
                    df_show = raw_pct_df.copy()
                    df_show.insert(0, "Posición", range(1, len(df_show) + 1))
                    df_show["Valor Mes Anterior (miles HNL)"] = (df_show["value_prev"] / 1000).round(0).apply(
                        lambda v: f"{v:,.0f}"
                    )
                    df_show["Valor Mes Actual (miles HNL)"] = (df_show["value_cur"] / 1000).round(0).apply(
                        lambda v: f"{v:,.0f}"
                    )
                    df_show["Variación Mensual %"] = df_show["pct_change"].apply(lambda v: f"{v:+.1f}%")
                    st.dataframe(
                        df_show[[
                            "Posición", "coop", "variable", "Dirección", "Variación Mensual %",
                            "Valor Mes Anterior (miles HNL)", "Valor Mes Actual (miles HNL)",
                        ]].rename(columns={
                            "coop": "Cooperativa", "variable": "Variable",
                        }),
                        use_container_width=True, hide_index=True,
                    )

                    # Todas las cuentas de la cooperativa seleccionada
                    st.markdown("##### 🔍 Detalle de la Cooperativa Seleccionada")
                    sel_coop_r = st.selectbox(
                        "Seleccione cooperativa",
                        df_show["coop"].drop_duplicates().tolist(),
                        key="rawpct_sel",
                    )
                    st.markdown("**📋 Datos brutos de los últimos 3 meses (todas las cuentas, miles HNL)**")
                    all_raw_r = get_recent_all_raw_data(
                        df_raw_all_clean, sel_coop_r, target_date, n_months=3,
                    )
                    if not all_raw_r.empty:
                        st.dataframe(all_raw_r, use_container_width=True, hide_index=True, height=420)
                    else:
                        st.info("No se encontraron datos brutos de esta cooperativa.")

                    # CSV
                    st.download_button(
                        "📥 Descargar CSV de Variación Mensual % (variables brutas)",
                        data=raw_pct_df.to_csv(index=False).encode("utf-8-sig"),
                        file_name=f"anomaly_raw_pct_{target_date:%Y%m}.csv",
                        mime="text/csv",
                    )


# =========================================================
# 13.6 🕸️ F-04 Descubrimiento Causal LiNGAM (agrupación anual)
# =========================================================
with tab_lingam:
    if df_ind.empty:
        st.error("No hay datos suficientes para calcular los indicadores financieros.")
    elif not LINGAM_OK:
        st.error("La librería `lingam` no está instalada. Ejecute `pip install lingam`.")
    else:
        st.markdown("#### 🕸️ Descubrimiento Causal LiNGAM (método de agrupación anual)")
        st.warning(
            "⚠️ **Importante**: LiNGAM infiere causalidad a partir de datos observacionales "
            "y sus resultados deben tratarse siempre como **hipótesis**. "
            "Verifique con experimentos de intervención o revisión de expertos."
        )
        st.info(
            "ℹ️ **Método de cálculo**: Es un **método anual** que, aunque "
            "técnicamente viola la independencia, prioriza la utilidad práctica. "
            "Se agrupan los 12 meses del año objetivo (menos si hay menos meses, "
            "como en el año 2026) en una única matriz de diseño cooperativa × mes "
            "y se ejecuta DirectLiNGAM sobre ella.\n\n"
            "- Como una misma cooperativa aparece varios meses, estrictamente "
            "cada muestra no es independiente.\n"
            "- Sin embargo, se asegura un tamaño de muestra mayor (hasta 12× el "
            "número de cooperativas) y el ruido mensual se promedia, lo que da "
            "resultados más estables (ventaja práctica).\n"
            "- Al cambiar de año puede compararse también el 'cambio estructural anual'."
        )
        st.caption(
            "Estructura: dos niveles — '① A nivel de indicadores (5 fijos)' y '② A nivel de variables brutas (selección libre)'"
        )

        # Subpestañas
        sub_lingam_ind, sub_lingam_raw = st.tabs([
            "① A nivel de indicadores (5 fijos)",
            "② A nivel de variables brutas (selección libre)",
        ])

        # === ① LiNGAM a nivel de indicadores ===
        with sub_lingam_ind:
            years_ind = available_years_from_ind(df_ind)
            if not years_ind:
                st.warning("No hay años disponibles en los datos.")
            else:
                default_y = default_lingam_year(years_ind, preferred=DEFAULT_LINGAM_YEAR)
                years_desc = sorted(years_ind, reverse=True)
                default_idx = years_desc.index(default_y) if default_y in years_desc else 0

                col_p1, col_p2 = st.columns(2)
                with col_p1:
                    target_year = st.selectbox(
                        "Año objetivo",
                        options=years_desc,
                        index=default_idx,
                        format_func=lambda y: f"Año {y}",
                        key="lingam_year",
                        help=f"Predeterminado: Año {DEFAULT_LINGAM_YEAR} (o el más reciente si no está)",
                    )
                with col_p2:
                    threshold = st.slider(
                        "Umbral de coeficiente causal a mostrar",
                        0.05, 0.30, 0.15, 0.05,
                        help="Solo se muestran flechas de coeficientes con valor absoluto mayor o igual",
                    )

                with st.spinner(f"Calculando LiNGAM (año {target_year}, agrupación anual)..."):
                    adj, labels, X_used, meta = run_lingam_year(df_ind, int(target_year))

                st.caption(
                    f"Año objetivo: **{meta['year']}** / "
                    f"Muestras: {meta['n_samples']:,} = "
                    f"{meta['n_coops']} cooperativas × {meta['n_months']} meses / "
                    f"Meses objetivo: {', '.join(meta['months']) if meta['months'] else '—'}"
                )

                if adj is None:
                    st.warning(
                        "Datos insuficientes para ejecutar LiNGAM "
                        "(tras la agrupación anual se requieren al menos 10 muestras)."
                    )
                else:
                    # Grafo causal
                    if GRAPHVIZ_OK:
                        dot_l = lingam_to_graphviz(adj, labels, threshold=threshold)
                        if dot_l is not None:
                            st.graphviz_chart(dot_l, use_container_width=True)
                        else:
                            fig_l = lingam_to_plotly(adj, labels, threshold=threshold)
                            st.plotly_chart(fig_l, use_container_width=True, key="f04_lingam")
                    else:
                        st.warning(
                            "La librería `graphviz` no está instalada. Se usa la versión en Plotly. "
                            "Se recomienda instalar `pip install graphviz` además del binario de Graphviz."
                        )
                        fig_l = lingam_to_plotly(adj, labels, threshold=threshold)
                        st.plotly_chart(fig_l, use_container_width=True, key="f04_lingam")
                    st.caption(
                        "🔵 Flecha azul: coeficiente positivo (al subir la causa, sube el efecto) / "
                        "🔴 Flecha roja: coeficiente negativo (al subir la causa, baja el efecto) / "
                        "Etiqueta = coeficiente, grosor = |coeficiente|"
                    )

                    # Diagnóstico de no gausianidad (agrupación anual)
                    st.markdown("##### Diagnóstico de no gausianidad de Shapiro-Wilk (agrupación anual)")
                    sw = shapiro_wilk_table_year(df_ind, int(target_year))
                    if not sw.empty:
                        valid = sw.dropna(subset=["No Gausiano"])
                        n_total = len(valid)
                        n_ng = (valid["No Gausiano"] == "✓").sum()
                        st.markdown(
                            f"**No Gausianos: {n_ng} / {n_total} indicadores** "
                            f"(Shapiro-Wilk p<0.05 se considera no gausiano)"
                        )
                        if n_total > 0:
                            if n_ng / n_total >= 0.7:
                                st.success("✅ El supuesto de no gausianidad de LiNGAM se cumple bien.")
                            elif n_ng / n_total >= 0.4:
                                st.warning("⚠️ La interpretación de los resultados requiere precaución.")
                            else:
                                st.error("❌ Los datos están próximos a una distribución gausiana; "
                                         "trate los resultados de LiNGAM solo como referencia.")
                        st.dataframe(sw, use_container_width=True, hide_index=True)

                    # Matriz de adyacencia
                    with st.expander("Valores numéricos: Matriz de Adyacencia (referencia)"):
                        df_adj = pd.DataFrame(adj, index=labels, columns=labels).round(3)
                        st.caption("Filas = efecto; Columnas = causa. Valor = coeficiente causal (estandarizado) que la columna da a la fila.")
                        st.dataframe(df_adj, use_container_width=True)

        # === ② LiNGAM a nivel de variables brutas ===
        with sub_lingam_raw:
            st.info(
                "💡 **Descubrimiento causal con variables brutas**. "
                "Se puede seleccionar libremente cualquier variable. "
                "Se agrupan en cooperativa × mes del año objetivo y se ejecuta DirectLiNGAM."
            )
            st.caption(
                "ℹ️ Dado que las cooperativas tienen tamaños (activo total) muy variables, "
                "el LiNGAM con valores absolutos (HNL) corre el riesgo de confundir la "
                "covariación típica ('las cooperativas grandes son grandes en todo') con "
                "causalidad. Lea los resultados **como hipótesis** y compárelos con el "
                "LiNGAM a nivel de indicadores (en proporciones)."
            )
            if df_raw_all_clean.empty:
                st.warning("No se cargaron datos brutos.")
            else:
                all_vars = sorted(df_raw_all_clean["variable"].dropna().unique().tolist())

                # Variables predeterminadas (resolución por coincidencia parcial)
                RAW_LINGAM_DEFAULT_PATTERNS = [
                    "GASTOS DE ADMINISTRACION",
                    "GASTOS FINANCIEROS",
                    "PRODUCTOS FINANCIEROS",
                    "PRODUCTOS POR SERVICIOS",
                    "INVERSIONES",
                    "PATRIMONIO",
                ]

                def _resolve_default(patterns: list[str], pool: list[str]) -> list[str]:
                    chosen: list[str] = []
                    for pat in patterns:
                        pat_norm = pat.strip().upper()
                        if pat_norm in pool and pat_norm not in chosen:
                            chosen.append(pat_norm)
                            continue
                        for v in pool:
                            if v == pat_norm:
                                if v not in chosen:
                                    chosen.append(v)
                                break
                        else:
                            for v in pool:
                                if v.startswith(pat_norm) and v not in chosen:
                                    chosen.append(v)
                                    break
                    return chosen

                default_vars = _resolve_default(RAW_LINGAM_DEFAULT_PATTERNS, all_vars)

                years_raw = available_years_from_raw(df_raw_all_clean)
                if not years_raw:
                    st.warning("No se pudieron extraer años de los datos brutos.")
                else:
                    default_y_r = default_lingam_year(
                        years_raw, preferred=DEFAULT_LINGAM_YEAR,
                    )
                    years_raw_desc = sorted(years_raw, reverse=True)
                    default_idx_r = (
                        years_raw_desc.index(default_y_r)
                        if default_y_r in years_raw_desc else 0
                    )

                    col_rp1, col_rp2 = st.columns(2)
                    with col_rp1:
                        target_year_raw = st.selectbox(
                            "Año objetivo",
                            options=years_raw_desc,
                            index=default_idx_r,
                            format_func=lambda y: f"Año {y}",
                            key="lingam_raw_year",
                            help=f"Predeterminado: Año {DEFAULT_LINGAM_YEAR} (o el más reciente si no está)",
                        )
                    with col_rp2:
                        threshold_raw = st.slider(
                            "Umbral de coeficiente causal a mostrar",
                            0.05, 0.30, 0.15, 0.05,
                            key="lingam_raw_thr",
                            help="Solo se muestran flechas de coeficientes con valor absoluto mayor o igual",
                        )

                    selected_vars = st.multiselect(
                        "Variables brutas a usar en LiNGAM (al menos 2)",
                        options=all_vars,
                        default=default_vars,
                        key="lingam_raw_vars",
                    )

                    if len(selected_vars) < 2:
                        st.warning("Seleccione al menos 2 variables.")
                    else:
                        with st.spinner(
                            f"Calculando LiNGAM de variables brutas (año {target_year_raw}, agrupación anual)..."
                        ):
                            adj_r, labels_r, X_used_r, meta_r = run_lingam_raw_year(
                                df_raw_all_clean, selected_vars, int(target_year_raw),
                            )

                        st.caption(
                            f"Año objetivo: **{meta_r['year']}** / "
                            f"Muestras: {meta_r['n_samples']:,} = "
                            f"{meta_r['n_coops']} cooperativas × {meta_r['n_months']} meses / "
                            f"Meses objetivo: {', '.join(meta_r['months']) if meta_r['months'] else '—'}"
                        )

                        if adj_r is None:
                            st.warning(
                                "Datos insuficientes para ejecutar LiNGAM de variables brutas "
                                "(se requieren al menos 10 muestras con todas las variables "
                                "seleccionadas disponibles en el año objetivo)."
                            )
                        else:
                            if GRAPHVIZ_OK:
                                dot_lr = lingam_to_graphviz(
                                    adj_r, labels_r, threshold=threshold_raw,
                                )
                                if dot_lr is not None:
                                    st.graphviz_chart(dot_lr, use_container_width=True)
                                else:
                                    fig_lr = lingam_to_plotly(
                                        adj_r, labels_r, threshold=threshold_raw,
                                    )
                                    st.plotly_chart(
                                        fig_lr, use_container_width=True,
                                        key="f04_lingam_raw",
                                    )
                            else:
                                st.warning(
                                    "La librería `graphviz` no está instalada. Se usa la versión en Plotly. "
                                    "Se recomienda instalar `pip install graphviz` además del binario de Graphviz."
                                )
                                fig_lr = lingam_to_plotly(
                                    adj_r, labels_r, threshold=threshold_raw,
                                )
                                st.plotly_chart(
                                    fig_lr, use_container_width=True,
                                    key="f04_lingam_raw",
                                )
                            st.caption(
                                "🔵 Flecha azul: coeficiente positivo / 🔴 Flecha roja: coeficiente negativo / "
                                "Etiqueta = coeficiente, grosor = |coeficiente| / "
                                f"Muestras usadas: {len(X_used_r):,}"
                            )

                            # Diagnóstico de no gausianidad (variables brutas, agrupación anual)
                            st.markdown("##### Diagnóstico de no gausianidad de Shapiro-Wilk (variables brutas, agrupación anual)")
                            sw_rows = []
                            for col in labels_r:
                                x = X_used_r[col].dropna().values
                                n = len(x)
                                if n < 20:
                                    sw_rows.append({
                                        "Variable": col, "n": n,
                                        "p-valor Shapiro-Wilk": None, "No Gausiano": None,
                                    })
                                    continue
                                try:
                                    _, p = stats.shapiro(x)
                                    sw_rows.append({
                                        "Variable": col, "n": n,
                                        "p-valor Shapiro-Wilk": round(float(p), 4),
                                        "No Gausiano": "✓" if p < 0.05 else "—",
                                    })
                                except Exception:
                                    sw_rows.append({
                                        "Variable": col, "n": n,
                                        "p-valor Shapiro-Wilk": None, "No Gausiano": None,
                                    })
                            sw_r = pd.DataFrame(sw_rows)
                            valid_r = sw_r.dropna(subset=["No Gausiano"])
                            n_total_r = len(valid_r)
                            n_ng_r = (valid_r["No Gausiano"] == "✓").sum()
                            st.markdown(
                                f"**No Gausianos: {n_ng_r} / {n_total_r} variables** "
                                f"(Shapiro-Wilk p<0.05 se considera no gausiano)"
                            )
                            if n_total_r > 0:
                                ratio = n_ng_r / n_total_r
                                if ratio >= 0.7:
                                    st.success("✅ El supuesto de no gausianidad de LiNGAM se cumple bien.")
                                elif ratio >= 0.4:
                                    st.warning("⚠️ La interpretación de los resultados requiere precaución.")
                                else:
                                    st.error("❌ Los datos están próximos a una distribución gausiana; "
                                             "trate los resultados de LiNGAM solo como referencia.")
                            st.dataframe(sw_r, use_container_width=True, hide_index=True)

                            with st.expander("Valores numéricos: Matriz de Adyacencia (referencia)"):
                                df_adj_r = pd.DataFrame(
                                    adj_r, index=labels_r, columns=labels_r,
                                ).round(3)
                                st.caption(
                                    "Filas = efecto; Columnas = causa. "
                                    "Valor = coeficiente causal (estandarizado) que la columna da a la fila."
                                )
                                st.dataframe(df_adj_r, use_container_width=True)


# ---- Pie de página ----
st.divider()
st.caption(
    "**EF v16 — Análisis de Estados Financieros de CONSUCOOP** / "
    f"4 módulos (F-01 a F-04) + Tablero / LLM = `{LLM_MODEL}` local / "
    "Representativo = mediana / LiNGAM = método de agrupación anual (predeterminado: 2025)"
)
